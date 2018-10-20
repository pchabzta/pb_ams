from .forms import PFormRM101A, PFormRM102A, PFormRM103A, PFormRM104A, PFormRM105A, PFormRM106A
from .forms import PFormRM201A, PFormRM202A, PFormRM203A, PFormRM204A, PFormRM205A, PFormRM206A
from .forms import PFormRM301A, PFormRM302A, PFormRM303A, PFormRM304A, PFormRM305A, PFormRM306A
from .forms import PFormRM201B, PFormRM202B, PFormRM203B, PFormRM204B, PFormRM205B
from .forms import PFormRM301B, PFormRM302B, PFormRM303B, PFormRM304B, PFormRM305B
from .forms import PFormRM401B, PFormRM402B, PFormRM403B, PFormRM404B, PFormRM405B
from .forms import PaymentForm
from .forms import Elec_cpu_change, Water_cpu_change, PhoneNoMessage
from .forms import MaintenanceForm
from .models import Extra, Room, Room_type, MaintenanceCharge, TenantProfile
import datetime
from django.utils.dateparse import parse_datetime, parse_date
from django.utils.timezone import is_aware, is_naive, make_aware, make_naive
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from .forms import TenantCreateForm, TenantProfileCreateForm
from django.contrib import messages
from django.http import HttpResponseRedirect
from django.urls import reverse_lazy
from django.shortcuts import get_object_or_404
from .forms import RM101A_BillForm, RM102A_BillForm, RM103A_BillForm, RM104A_BillForm, RM105A_BillForm, RM106A_BillForm
from .forms import RM201A_BillForm, RM202A_BillForm, RM203A_BillForm, RM204A_BillForm, RM205A_BillForm, RM206A_BillForm
from .forms import RM301A_BillForm, RM302A_BillForm, RM303A_BillForm, RM304A_BillForm, RM305A_BillForm, RM306A_BillForm
from .forms import RM201B_BillForm, RM202B_BillForm, RM203B_BillForm, RM204B_BillForm, RM205B_BillForm
from .forms import RM301B_BillForm, RM302B_BillForm, RM303B_BillForm, RM304B_BillForm, RM305B_BillForm
from .forms import RM401B_BillForm, RM402B_BillForm, RM403B_BillForm, RM404B_BillForm, RM405B_BillForm
from ams.models import Billing
import random
import calendar

import decimal

import GV

import os
from openpyxl import workbook, load_workbook

# -----SMS FROM LOCALHOST & WEB------------------------------------------
import os
from twilio.rest import Client
from twilio.http.http_client import TwilioHttpClient
# -----------------------------------------------------------------------
import GV


@login_required
def gateway(request):
    if str(request.user) in ['Admin Admin', 'Preecha Bootwicha']:

        return HttpResponseRedirect(reverse_lazy('admin_page'))
    else:
        return HttpResponseRedirect(reverse_lazy('tenant_page'))


@login_required
def create_contract(request):
    if request.method == 'POST':

        tenant_form = TenantCreateForm(data=request.POST)
        # tenant_profile_form = TenantProfileCreateForm(data=request.POST, files=request.FILES)
        tenant_profile_form = TenantProfileCreateForm(data=request.POST, files=request.FILES)

        if tenant_form.is_valid() and tenant_profile_form.is_valid():

            # Create a new tenant object but avoid saving it yet
            new_tenant = tenant_form.save(commit=False)

            # Set the chosen password
            # new_tenant.set_password(tenant_form.cleaned_data['password'])
            new_tenant.set_password(tenant_form.clean_password2())

            # Save the new_tenant object
            new_tenant.save()

            # Create a new tenantprofile object but avoid saving it yet
            tenant_profile = tenant_profile_form.save(commit=False)  # save_m2m() added to tenant_profile_form

            # Set the chosen tenant field
            tenant_profile.tenant = new_tenant

            # ------------------------------------------
            # provide initial value to certain fields before saving to DB
            tenant_profile.elec_unit = 0
            tenant_profile.water_unit = 0
            tenant_profile.misc_cost = 0
            # -----------------------------------------

            # Save the tenantprofile object
            tenant_profile.save()

            # Save the ManyToMany
            tenant_profile_form.save_m2m()

            messages.success(request, 'Profile has been updated successfully')

            return HttpResponseRedirect(reverse_lazy('admin_page'))
        else:
            messages.error(request, 'Error updating your tenant_profile')

    else:
        tenant_form = TenantCreateForm()
        # tenant_profile_form = TenantProfileCreateForm()
        tenant_profile_form = TenantProfileCreateForm()

    return render(request, 'ams/create_contract.html',
                  {'section': 'new_contract', 'tenant_form': tenant_form, 'tenant_profile_form': tenant_profile_form})


# @login_required # ????? #ORIGINAL
def create_bill(room_no):
    pf = get_object_or_404(TenantProfile, room_no__room_no=room_no)
    tname = pf.tenant.first_name + ' ' + pf.tenant.last_name

    rno = pf.room_no.room_no
    adj = pf.adjust

    exd = {}
    exd.setdefault('Electricity CPU', 0)
    exd.setdefault('Water CPU', 0)
    exd.setdefault('Garbage', 0)
    exd.setdefault('Parking', 0)
    exd.setdefault('Wifi', 0)
    exd.setdefault('Cable TV', 0)
    exd.setdefault('Bed', 0)
    exd.setdefault('Bed accessories', 0)
    exd.setdefault('Dressing Table', 0)
    exd.setdefault('Clothing Cupboard', 0)
    exd.setdefault('TV Table', 0)
    exd.setdefault('Fridge', 0)
    exd.setdefault('Air-Conditioner', 0)

    for e in pf.extra.all():
        exd.update({e.desc: e.cpu})

    room_cost = pf.room_no.room_type.rate
    room_acc_cost = exd['Bed'] + exd['Bed accessories'] + exd['Dressing Table'] \
                    + exd['Clothing Cupboard'] + exd['TV Table'] + exd['Fridge'] \
                    + exd['Air-Conditioner']

    elec_cost = exd['Electricity CPU'] * pf.elec_unit
    water_cost = exd['Water CPU'] * pf.water_unit

    com_ser_cost = pf.elec_unit * GV.COMMOM_SERVICE_CPU

    oth_ser_cost = exd['Garbage'] + exd['Parking'] + exd['Wifi'] + exd['Cable TV']
    ovd_amt = pf.cum_ovd

    # -----------------------
    late_f = pf.late_fee
    maint_c = pf.maint_cost

    # RESET pf.late_fee & pf.maint_cost TO O TO BE READY FOR NEXT CYCLE
    pf.late_fee = 0
    pf.maint_cost = 0
    # -----------------------

    total = room_cost + room_acc_cost + elec_cost + water_cost + com_ser_cost + oth_ser_cost + ovd_amt + adj + late_f + maint_c

    # CREATE PRELIMINARY BILL OBJECT **************
    new_bill = Billing(bill_ref=get_ref_string(),
                       bill_date=datetime.datetime.now().date(),  # SUPPLY BILL DATE
                       tenant_name=tname,
                       room_no=rno,
                       room_cost=room_cost,
                       room_acc_cost=room_acc_cost,
                       electricity_cost=elec_cost,
                       water_cost=water_cost,
                       common_ser_cost=com_ser_cost,
                       other_ser_cost=oth_ser_cost,
                       overdue_amount=ovd_amt,

                       # -----------------------
                       late_fee=late_f,
                       maint_cost=maint_c,
                       # -----------------------

                       adjust=adj,
                       bill_total=total,

                       )

    # SAVE TENANTPROFILE OBJECT TO DB
    pf.save()

    # ADJUST PRELIMINARY BILL OBJECT
    adjust_bill(pf, new_bill)


def adjust_bill(pf, new_bill):
    tn_bill = new_bill

    bref = tn_bill.bill_ref
    bdate = tn_bill.bill_date
    # bupd # TO BE FILLED WHEN SAVED
    # bstat # TO BE FILLED WHEN SAVED
    tname = tn_bill.tenant_name
    rno = tn_bill.room_no
    room_cost = tn_bill.room_cost
    room_acc_cost = tn_bill.room_acc_cost
    elec_cost = tn_bill.electricity_cost
    water_cost = tn_bill.water_cost
    com_ser_cost = tn_bill.common_ser_cost
    oth_ser_cost = tn_bill.other_ser_cost
    ovd_amt = tn_bill.overdue_amount
    adj = tn_bill.adjust
    # total = tn_bill.bill_total # TO BE ADJUSTED IF REQUIRED

    # pay_date # TO BE FILLED AT PAYMENT
    # pay_amt #TO BE FILL AT PAYMENT
    # bf #TO BE FILLED AT PAYMENT

    late_f = tn_bill.late_fee
    maint_c = tn_bill.maint_cost

    sdate = pf.start_date  # FROM pf

    start_day = sdate.day
    bill_day = bdate.day

    start_m = sdate.month
    bill_m = bdate.month

    number_of_day_in_start_month = calendar.monthrange(sdate.year, sdate.month)[1]
    nodsm = number_of_day_in_start_month
    number_of_day_in_bill_month = calendar.monthrange(bdate.year, bdate.month)[1]
    nodbm = number_of_day_in_bill_month

    if abs(start_m - bill_m) == 0:
        tbd = number_of_day_in_bill_month - start_day + 1  # SPECIAL CASE 1
    elif abs(start_m - bill_m) == 1 and start_day >= bill_day:
        tbd = number_of_day_in_bill_month + (number_of_day_in_start_month - start_day + 1)  # SPECIAL CASE 2
    else:
        tbd = number_of_day_in_bill_month  # ONGOING CASE

    # ADJUST CERTAIN VALUES IN PRELIM. BILL OBJECT
    const = decimal.Decimal((tbd / nodbm))

    room_cost = room_cost * const
    room_acc_cost = room_acc_cost * const
    com_ser_cost = com_ser_cost * const
    oth_ser_cost = oth_ser_cost * const
    adj = adj * const

    total = (room_cost + room_acc_cost + adj) + elec_cost + water_cost + (
            com_ser_cost + oth_ser_cost) + ovd_amt + late_f + maint_c

    # CREATE FINAL BILL OBJECT *******************
    new_bill = Billing(bill_ref=bref,
                       tenant_name=tname,
                       room_no=rno,
                       room_cost=room_cost,
                       room_acc_cost=room_acc_cost,
                       electricity_cost=elec_cost,
                       water_cost=water_cost,
                       common_ser_cost=com_ser_cost,
                       other_ser_cost=oth_ser_cost,
                       overdue_amount=ovd_amt,

                       # -----------------------
                       late_fee=late_f,
                       maint_cost=maint_c,
                       # -----------------------

                       adjust=adj,
                       bill_total=total,

                       )

    # SAVE BILL OBJECT TO DB
    new_bill.save()


@login_required
def billing(request):
    tenant_pf = TenantProfile.objects.order_by("room_no")

    rm101a_form = None
    rm102a_form = None
    rm103a_form = None
    rm104a_form = None
    rm105a_form = None
    rm106a_form = None

    rm201a_form = None
    rm202a_form = None
    rm203a_form = None
    rm204a_form = None
    rm205a_form = None
    rm206a_form = None

    rm301a_form = None
    rm302a_form = None
    rm303a_form = None
    rm304a_form = None
    rm305a_form = None
    rm306a_form = None

    rm201b_form = None
    rm202b_form = None
    rm203b_form = None
    rm204b_form = None
    rm205b_form = None

    rm301b_form = None
    rm302b_form = None
    rm303b_form = None
    rm304b_form = None
    rm305b_form = None

    rm401b_form = None
    rm402b_form = None
    rm403b_form = None
    rm404b_form = None
    rm405b_form = None

    no_of_bill = 0
    for tpf in tenant_pf:
        rmn = tpf.room_no.room_no

        if request.method == 'POST':

            if rmn == '101A':
                pf = get_object_or_404(TenantProfile, room_no__room_no=rmn)
                rm101a_form = RM101A_BillForm(data=request.POST, instance=pf, prefix='rm101a')
                if rm101a_form.is_valid():
                    rm101a_form.save(commit=True)
                    # -------------------
                    create_bill(rmn)

                    no_of_bill += 1
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 101A Billing')
            if rmn == '102A':
                pf = get_object_or_404(TenantProfile, room_no__room_no=rmn)
                rm102a_form = RM102A_BillForm(data=request.POST, instance=pf, prefix='rm102a')
                if rm102a_form.is_valid():
                    rm102a_form.save(commit=True)
                    # -------------------
                    create_bill(rmn)
                    no_of_bill += 1
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 102A Billing')
            if rmn == '103A':
                pf = get_object_or_404(TenantProfile, room_no__room_no=rmn)
                rm103a_form = RM103A_BillForm(data=request.POST, instance=pf, prefix='rm103a')
                if rm103a_form.is_valid():
                    rm103a_form.save(commit=True)
                    # -------------------
                    create_bill(rmn)
                    no_of_bill += 1
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 103A Billing')

            if rmn == '104A':
                pf = get_object_or_404(TenantProfile, room_no__room_no=rmn)
                rm104a_form = RM104A_BillForm(data=request.POST, instance=pf, prefix='rm104a')
                if rm104a_form.is_valid():
                    rm104a_form.save(commit=True)
                    # -------------------
                    create_bill(rmn)
                    no_of_bill += 1
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 104A Billing')

            if rmn == '105A':
                pf = get_object_or_404(TenantProfile, room_no__room_no=rmn)
                rm105a_form = RM105A_BillForm(data=request.POST, instance=pf, prefix='rm105a')
                if rm105a_form.is_valid():
                    rm105a_form.save(commit=True)
                    # -------------------
                    create_bill(rmn)
                    no_of_bill += 1
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 105A Billing')

            if rmn == '106A':
                pf = get_object_or_404(TenantProfile, room_no__room_no=rmn)
                rm106a_form = RM106A_BillForm(data=request.POST, instance=pf, prefix='rm106a')
                if rm106a_form.is_valid():
                    rm106a_form.save(commit=True)
                    # -------------------
                    create_bill(rmn)
                    no_of_bill += 1
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 106A Billing')

            if rmn == '201A':
                pf = get_object_or_404(TenantProfile, room_no__room_no=rmn)
                rm201a_form = RM201A_BillForm(data=request.POST, instance=pf, prefix='rm201a')
                if rm201a_form.is_valid():
                    rm201a_form.save(commit=True)
                    # -------------------
                    create_bill(rmn)
                    no_of_bill += 1
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 201A Billing')

            if rmn == '202A':
                pf = get_object_or_404(TenantProfile, room_no__room_no=rmn)
                rm202a_form = RM202A_BillForm(data=request.POST, instance=pf, prefix='rm202a')
                if rm202a_form.is_valid():
                    rm202a_form.save(commit=True)
                    # -------------------
                    create_bill(rmn)
                    no_of_bill += 1
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 202A Billing')

            if rmn == '203A':
                pf = get_object_or_404(TenantProfile, room_no__room_no=rmn)
                rm203a_form = RM203A_BillForm(data=request.POST, instance=pf, prefix='rm203a')
                if rm203a_form.is_valid():
                    rm203a_form.save(commit=True)
                    # -------------------
                    create_bill(rmn)
                    no_of_bill += 1
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 203A Billing')

            if rmn == '204A':
                pf = get_object_or_404(TenantProfile, room_no__room_no=rmn)
                rm204a_form = RM204A_BillForm(data=request.POST, instance=pf, prefix='rm204a')
                if rm204a_form.is_valid():
                    rm204a_form.save(commit=True)
                    # -------------------
                    create_bill(rmn)
                    no_of_bill += 1
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 204A Billing')

            if rmn == '205A':
                pf = get_object_or_404(TenantProfile, room_no__room_no=rmn)
                rm205a_form = RM205A_BillForm(data=request.POST, instance=pf, prefix='rm205a')
                if rm205a_form.is_valid():
                    rm205a_form.save(commit=True)
                    # -------------------
                    create_bill(rmn)
                    no_of_bill += 1
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 205A Billing')

            if rmn == '206A':
                pf = get_object_or_404(TenantProfile, room_no__room_no=rmn)
                rm206a_form = RM206A_BillForm(data=request.POST, instance=pf, prefix='rm206a')
                if rm206a_form.is_valid():
                    rm206a_form.save(commit=True)
                    # -------------------
                    create_bill(rmn)
                    no_of_bill += 1
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 206A Billing')

            if rmn == '301A':
                pf = get_object_or_404(TenantProfile, room_no__room_no=rmn)
                rm301a_form = RM301A_BillForm(data=request.POST, instance=pf, prefix='rm301a')
                if rm301a_form.is_valid():
                    rm301a_form.save(commit=True)
                    # -------------------
                    create_bill(rmn)
                    no_of_bill += 1
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 301A Billing')

            if rmn == '302A':
                pf = get_object_or_404(TenantProfile, room_no__room_no=rmn)
                rm302a_form = RM302A_BillForm(data=request.POST, instance=pf, prefix='rm302a')
                if rm302a_form.is_valid():
                    rm302a_form.save(commit=True)
                    # -------------------
                    create_bill(rmn)
                    no_of_bill += 1
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 302A Billing')

            if rmn == '303A':
                pf = get_object_or_404(TenantProfile, room_no__room_no=rmn)
                rm303a_form = RM303A_BillForm(data=request.POST, instance=pf, prefix='rm303a')
                if rm303a_form.is_valid():
                    rm303a_form.save(commit=True)
                    # -------------------
                    create_bill(rmn)
                    no_of_bill += 1
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 303A Billing')

            if rmn == '304A':
                pf = get_object_or_404(TenantProfile, room_no__room_no=rmn)
                rm304a_form = RM304A_BillForm(data=request.POST, instance=pf, prefix='rm304a')
                if rm304a_form.is_valid():
                    rm304a_form.save(commit=True)
                    # -------------------
                    create_bill(rmn)
                    no_of_bill += 1
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 304A Billing')

            if rmn == '305A':
                pf = get_object_or_404(TenantProfile, room_no__room_no=rmn)
                rm305a_form = RM305A_BillForm(data=request.POST, instance=pf, prefix='rm305a')
                if rm305a_form.is_valid():
                    rm305a_form.save(commit=True)
                    # -------------------
                    create_bill(rmn)
                    no_of_bill += 1
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 305A Billing')

            if rmn == '306A':
                pf = get_object_or_404(TenantProfile, room_no__room_no=rmn)
                rm306a_form = RM306A_BillForm(data=request.POST, instance=pf, prefix='rm306a')
                if rm306a_form.is_valid():
                    rm306a_form.save(commit=True)
                    # -------------------
                    create_bill(rmn)
                    no_of_bill += 1
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 306A Billing')

            if rmn == '201B':
                pf = get_object_or_404(TenantProfile, room_no__room_no=rmn)
                rm201b_form = RM201B_BillForm(data=request.POST, instance=pf, prefix='rm201b')
                if rm201b_form.is_valid():
                    rm201b_form.save(commit=True)
                    # -------------------
                    create_bill(rmn)
                    no_of_bill += 1
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 201B Billing')

            if rmn == '202B':
                pf = get_object_or_404(TenantProfile, room_no__room_no=rmn)
                rm202b_form = RM202B_BillForm(data=request.POST, instance=pf, prefix='rm202b')
                if rm202b_form.is_valid():
                    rm202b_form.save(commit=True)
                    # -------------------
                    create_bill(rmn)
                    no_of_bill += 1
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 202B Billing')

            if rmn == '203B':
                pf = get_object_or_404(TenantProfile, room_no__room_no=rmn)
                rm203b_form = RM203B_BillForm(data=request.POST, instance=pf, prefix='rm203b')
                if rm203b_form.is_valid():
                    rm203b_form.save(commit=True)
                    # -------------------
                    create_bill(rmn)
                    no_of_bill += 1
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 203B Billing')

            if rmn == '204B':
                pf = get_object_or_404(TenantProfile, room_no__room_no=rmn)
                rm204b_form = RM204B_BillForm(data=request.POST, instance=pf, prefix='rm204b')
                if rm204b_form.is_valid():
                    rm204b_form.save(commit=True)
                    # -------------------
                    create_bill(rmn)
                    no_of_bill += 1
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 204B Billing')

            if rmn == '205B':
                pf = get_object_or_404(TenantProfile, room_no__room_no=rmn)
                rm205b_form = RM205B_BillForm(data=request.POST, instance=pf, prefix='rm205b')
                if rm205b_form.is_valid():
                    rm205b_form.save(commit=True)
                    # -------------------
                    create_bill(rmn)
                    no_of_bill += 1
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 205B Billing')

            if rmn == '301B':
                pf = get_object_or_404(TenantProfile, room_no__room_no=rmn)
                rm301b_form = RM301B_BillForm(data=request.POST, instance=pf, prefix='rm301b')
                if rm301b_form.is_valid():
                    rm301b_form.save(commit=True)
                    # -------------------
                    create_bill(rmn)
                    no_of_bill += 1
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 301B Billing')

            if rmn == '302B':
                pf = get_object_or_404(TenantProfile, room_no__room_no=rmn)
                rm302b_form = RM302B_BillForm(data=request.POST, instance=pf, prefix='rm302b')
                if rm302b_form.is_valid():
                    rm302b_form.save(commit=True)
                    # -------------------
                    create_bill(rmn)
                    no_of_bill += 1
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 302B Billing')

            if rmn == '303B':
                pf = get_object_or_404(TenantProfile, room_no__room_no=rmn)
                rm303b_form = RM303B_BillForm(data=request.POST, instance=pf, prefix='rm303b')
                if rm303b_form.is_valid():
                    rm303b_form.save(commit=True)
                    # -------------------
                    create_bill(rmn)
                    no_of_bill += 1
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 303B Billing')

            if rmn == '304B':
                pf = get_object_or_404(TenantProfile, room_no__room_no=rmn)
                rm304b_form = RM304B_BillForm(data=request.POST, instance=pf, prefix='rm304b')
                if rm304b_form.is_valid():
                    rm304b_form.save(commit=True)
                    # -------------------
                    create_bill(rmn)
                    no_of_bill += 1
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 304B Billing')

            if rmn == '305B':
                pf = get_object_or_404(TenantProfile, room_no__room_no=rmn)
                rm305b_form = RM305B_BillForm(data=request.POST, instance=pf, prefix='rm305b')
                if rm305b_form.is_valid():
                    rm305b_form.save(commit=True)
                    # -------------------
                    create_bill(rmn)
                    no_of_bill += 1
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 305B Billing')

            if rmn == '401B':
                pf = get_object_or_404(TenantProfile, room_no__room_no=rmn)
                rm401b_form = RM401B_BillForm(data=request.POST, instance=pf, prefix='rm401b')
                if rm401b_form.is_valid():
                    rm401b_form.save(commit=True)
                    # -------------------
                    create_bill(rmn)
                    no_of_bill += 1
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 401B Billing')

            if rmn == '402B':
                pf = get_object_or_404(TenantProfile, room_no__room_no=rmn)
                rm402b_form = RM402B_BillForm(data=request.POST, instance=pf, prefix='rm402b')
                if rm402b_form.is_valid():
                    rm402b_form.save(commit=True)
                    # -------------------
                    create_bill(rmn)
                    no_of_bill += 1
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 402B Billing')

            if rmn == '403B':
                pf = get_object_or_404(TenantProfile, room_no__room_no=rmn)
                rm403b_form = RM403B_BillForm(data=request.POST, instance=pf, prefix='rm403b')
                if rm403b_form.is_valid():
                    rm403b_form.save(commit=True)
                    # -------------------
                    create_bill(rmn)
                    no_of_bill += 1
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 403B Billing')

            if rmn == '404B':
                pf = get_object_or_404(TenantProfile, room_no__room_no=rmn)
                rm404b_form = RM404B_BillForm(data=request.POST, instance=pf, prefix='rm404b')
                if rm404b_form.is_valid():
                    rm404b_form.save(commit=True)
                    # -------------------
                    create_bill(rmn)
                    no_of_bill += 1
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 404B Billing')

            if rmn == '405B':
                pf = get_object_or_404(TenantProfile, room_no__room_no=rmn)
                rm405b_form = RM405B_BillForm(data=request.POST, instance=pf, prefix='rm405b')
                if rm405b_form.is_valid():
                    rm405b_form.save(commit=True)
                    # -------------------
                    create_bill(rmn)
                    no_of_bill += 1
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 405B Billing')

        else:
            if rmn == '101A':
                rm101a_form = RM101A_BillForm(prefix='rm101a')

            if rmn == '102A':
                rm102a_form = RM102A_BillForm(prefix='rm102a')

            if rmn == '103A':
                rm103a_form = RM103A_BillForm(prefix='rm103a')

            if rmn == '104A':
                rm104a_form = RM104A_BillForm(prefix='rm104a')

            if rmn == '105A':
                rm105a_form = RM105A_BillForm(prefix='rm105a')

            if rmn == '106A':
                rm106a_form = RM106A_BillForm(prefix='rm106a')

            if rmn == '201A':
                rm201a_form = RM201A_BillForm(prefix='rm201a')

            if rmn == '202A':
                rm202a_form = RM202A_BillForm(prefix='rm202a')

            if rmn == '203A':
                rm203a_form = RM203A_BillForm(prefix='rm203a')

            if rmn == '204A':
                rm204a_form = RM204A_BillForm(prefix='rm204a')

            if rmn == '205A':
                rm205a_form = RM205A_BillForm(prefix='rm205a')

            if rmn == '206A':
                rm206a_form = RM206A_BillForm(prefix='rm206a')

            if rmn == '301A':
                rm301a_form = RM301A_BillForm(prefix='rm301a')

            if rmn == '302A':
                rm302a_form = RM302A_BillForm(prefix='rm302a')

            if rmn == '303A':
                rm303a_form = RM303A_BillForm(prefix='rm303a')

            if rmn == '304A':
                rm304a_form = RM304A_BillForm(prefix='rm304a')

            if rmn == '305A':
                rm305a_form = RM305A_BillForm(prefix='rm305a')

            if rmn == '306A':
                rm306a_form = RM306A_BillForm(prefix='rm306a')

            if rmn == '201B':
                rm201b_form = RM201B_BillForm(prefix='rm201b')

            if rmn == '202B':
                rm202b_form = RM202B_BillForm(prefix='rm202b')

            if rmn == '203B':
                rm203b_form = RM203B_BillForm(prefix='rm203b')

            if rmn == '204B':
                rm204b_form = RM204B_BillForm(prefix='rm204b')

            if rmn == '205B':
                rm205b_form = RM205B_BillForm(prefix='rm205b')

            if rmn == '301B':
                rm301b_form = RM301B_BillForm(prefix='rm301b')

            if rmn == '302B':
                rm302b_form = RM302B_BillForm(prefix='rm302b')

            if rmn == '303B':
                rm303b_form = RM303B_BillForm(prefix='rm303b')

            if rmn == '304B':
                rm304b_form = RM304B_BillForm(prefix='rm304b')

            if rmn == '305B':
                rm305b_form = RM305B_BillForm(prefix='rm305b')

            if rmn == '401B':
                rm401b_form = RM401B_BillForm(prefix='rm401b')

            if rmn == '402B':
                rm402b_form = RM402B_BillForm(prefix='rm402b')

            if rmn == '403B':
                rm403b_form = RM403B_BillForm(prefix='rm403b')

            if rmn == '404B':
                rm404b_form = RM404B_BillForm(prefix='rm404b')

            if rmn == '405B':
                rm405b_form = RM405B_BillForm(prefix='rm405b')

    if request.method == 'POST':

        # WRITE TO BILL SUMMARY AND BILL SLIP (Localhost only, at this time !!!!)
        # create_exel_sheet(request)

        # -----------------
        # FOR PYTHONANYWHERE HOST (uncomment the following line !!)
        messages.success(request, 'Total {} bills have been created.'.format(no_of_bill))
        # -----------------
        return HttpResponseRedirect(reverse_lazy('admin_page'))
    else:
        return render(request, 'ams/billing.html', {'tenant_pf': tenant_pf, 'section': 'billing',
                                                    'rm101a_form': rm101a_form,
                                                    'rm102a_form': rm102a_form,
                                                    'rm103a_form': rm103a_form,
                                                    'rm104a_form': rm104a_form,
                                                    'rm105a_form': rm105a_form,
                                                    'rm106a_form': rm106a_form,

                                                    'rm201a_form': rm201a_form,
                                                    'rm202a_form': rm202a_form,
                                                    'rm203a_form': rm203a_form,
                                                    'rm204a_form': rm204a_form,
                                                    'rm205a_form': rm205a_form,
                                                    'rm206a_form': rm206a_form,

                                                    'rm301a_form': rm301a_form,
                                                    'rm302a_form': rm302a_form,
                                                    'rm303a_form': rm303a_form,
                                                    'rm304a_form': rm304a_form,
                                                    'rm305a_form': rm305a_form,
                                                    'rm306a_form': rm306a_form,

                                                    'rm201b_form': rm201b_form,
                                                    'rm202b_form': rm202b_form,
                                                    'rm203b_form': rm203b_form,
                                                    'rm204b_form': rm204b_form,
                                                    'rm205b_form': rm205b_form,

                                                    'rm301b_form': rm301b_form,
                                                    'rm302b_form': rm302b_form,
                                                    'rm303b_form': rm303b_form,
                                                    'rm304b_form': rm304b_form,
                                                    'rm305b_form': rm305b_form,

                                                    'rm401b_form': rm401b_form,
                                                    'rm402b_form': rm402b_form,
                                                    'rm403b_form': rm403b_form,
                                                    'rm404b_form': rm404b_form,
                                                    'rm405b_form': rm405b_form,

                                                    })


# @login_required (cannot be used here !!!)
def update_pf_and_bill(roomno, cd):
    pf = get_object_or_404(TenantProfile, room_no__room_no=roomno)
    bill = get_object_or_404(Billing, room_no=roomno, status='open')

    cf = bill.bill_total - cd['payment_amount']
    bill.cf_amount = cf
    pf.cum_ovd = cf
    bill.payment_date = cd['payment_date']
    bill.payment_amount = cd['payment_amount']
    bill.status = 'close'

    # CALCULATE LATE-FEE COST TO UPDATE PF.LATE_FEE
    bill_month = bill.bill_date.month

    pay_month = bill.payment_date.month
    pay_day = bill.payment_date.day

    late_fee = 0

    if pay_month > bill_month:
        if pay_day > GV.LATE_DAY_MAX:
            late_fee = GV.LATE_FEE_PER_DAY * (pay_day - GV.LATE_DAY_MAX)

    # Update pf for next billing
    pf.late_fee = late_fee

    # Update DB
    bill.save()
    pf.save()


@login_required
def pay_rent(request, bref):
    tenant_bill = get_object_or_404(Billing, bill_ref=bref, status='open')
    rmn = tenant_bill.room_no

    if request.method == 'POST':
        pay_form = PaymentForm(data=request.POST)

        if pay_form.is_valid():
            cd = pay_form.cleaned_data

            # -------------------
            update_pf_and_bill(rmn, cd)
            # ------------------

        else:
            messages.error(request, 'Error updating Room {} Payment'.format(tenant_bill.room_no))

    else:
        pay_form = PaymentForm()

    if request.method == 'POST':
        messages.success(request, 'Room {}: Payment has been completed !!!'.format(rmn))
        return HttpResponseRedirect(reverse_lazy('payment_individual'))
    else:
        return render(request, 'ams/pay_rent.html', {'tenant_bill': tenant_bill, 'pay_form': pay_form})


@login_required
def payment_individual(request):
    bills = Billing.objects.filter(status='open').order_by('id')

    return render(request, 'ams/payment_individual.html', {'bills': bills, 'section': 'payment_individual'})


@login_required
def payment(request):
    # bills were created in order from first room to last room
    bills = Billing.objects.filter(status='open').order_by('id')

    rm101a_form = None
    rm102a_form = None
    rm103a_form = None
    rm104a_form = None
    rm105a_form = None
    rm106a_form = None

    rm201a_form = None
    rm202a_form = None
    rm203a_form = None
    rm204a_form = None
    rm205a_form = None
    rm206a_form = None

    rm301a_form = None
    rm302a_form = None
    rm303a_form = None
    rm304a_form = None
    rm305a_form = None
    rm306a_form = None

    rm201b_form = None
    rm202b_form = None
    rm203b_form = None
    rm204b_form = None
    rm205b_form = None

    rm301b_form = None
    rm302b_form = None
    rm303b_form = None
    rm304b_form = None
    rm305b_form = None

    rm401b_form = None
    rm402b_form = None
    rm403b_form = None
    rm404b_form = None
    rm405b_form = None

    for bill in bills:
        rmn = bill.room_no

        if request.method == 'POST':

            if rmn == '101A':
                rm101a_form = PFormRM101A(data=request.POST, prefix='rm101a')
                if rm101a_form.is_valid():
                    cd = rm101a_form.cleaned_data

                    # -------------------
                    update_pf_and_bill(rmn, cd)
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 101A Payment')
            if rmn == '102A':
                rm102a_form = PFormRM102A(data=request.POST, prefix='rm102a')
                if rm102a_form.is_valid():
                    cd = rm102a_form.cleaned_data

                    # -------------------
                    update_pf_and_bill(rmn, cd)
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 102A Payment')

            if rmn == '103A':
                rm103a_form = PFormRM103A(data=request.POST, prefix='rm103a')
                if rm103a_form.is_valid():
                    cd = rm103a_form.cleaned_data

                    # -------------------
                    update_pf_and_bill(rmn, cd)
                    # ------------------

                else:
                    messages.error(request, 'Error updating Room 103A Payment')

            if rmn == '104A':
                rm104a_form = PFormRM104A(data=request.POST, prefix='rm104a')
                if rm104a_form.is_valid():
                    cd = rm104a_form.cleaned_data

                    # -------------------
                    update_pf_and_bill(rmn, cd)
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 104A Payment')

            if rmn == '105A':
                rm105a_form = PFormRM105A(data=request.POST, prefix='rm105a')
                if rm105a_form.is_valid():
                    cd = rm105a_form.cleaned_data

                    # -------------------
                    update_pf_and_bill(rmn, cd)
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 105A Payment')

            if rmn == '106A':
                rm106a_form = PFormRM106A(data=request.POST, prefix='rm106a')
                if rm106a_form.is_valid():
                    cd = rm106a_form.cleaned_data

                    # -------------------
                    update_pf_and_bill(rmn, cd)
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 106A Payment')

            if rmn == '201A':
                rm201a_form = PFormRM201A(data=request.POST, prefix='rm201a')
                if rm201a_form.is_valid():
                    cd = rm201a_form.cleaned_data

                    # -------------------
                    update_pf_and_bill(rmn, cd)
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 201A Payment')

            if rmn == '202A':
                rm202a_form = PFormRM202A(data=request.POST, prefix='rm202a')
                if rm202a_form.is_valid():
                    cd = rm202a_form.cleaned_data

                    # -------------------
                    update_pf_and_bill(rmn, cd)
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 202A Payment')

            if rmn == '203A':
                rm203a_form = PFormRM203A(data=request.POST, prefix='rm203a')
                if rm203a_form.is_valid():
                    cd = rm203a_form.cleaned_data

                    # -------------------
                    update_pf_and_bill(rmn, cd)
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 203A Payment')

            if rmn == '204A':
                rm204a_form = PFormRM204A(data=request.POST, prefix='rm204a')
                if rm204a_form.is_valid():
                    cd = rm204a_form.cleaned_data

                    # -------------------
                    update_pf_and_bill(rmn, cd)
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 204A Payment')

            if rmn == '205A':
                rm205a_form = PFormRM205A(data=request.POST, prefix='rm205a')
                if rm205a_form.is_valid():
                    cd = rm205a_form.cleaned_data

                    # -------------------
                    update_pf_and_bill(rmn, cd)
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 205A Payment')

            if rmn == '206A':
                rm206a_form = PFormRM206A(data=request.POST, prefix='rm206a')
                if rm206a_form.is_valid():
                    cd = rm206a_form.cleaned_data

                    # -------------------
                    update_pf_and_bill(rmn, cd)
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 206A Payment')

            if rmn == '301A':
                rm301a_form = PFormRM301A(data=request.POST, prefix='rm301a')
                if rm301a_form.is_valid():
                    cd = rm301a_form.cleaned_data

                    # -------------------
                    update_pf_and_bill(rmn, cd)
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 301A Payment')

            if rmn == '302A':
                rm302a_form = PFormRM302A(data=request.POST, prefix='rm302a')
                if rm302a_form.is_valid():
                    cd = rm302a_form.cleaned_data

                    # -------------------
                    update_pf_and_bill(rmn, cd)
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 302A Payment')

            if rmn == '303A':
                rm303a_form = PFormRM303A(data=request.POST, prefix='rm303a')
                if rm303a_form.is_valid():
                    cd = rm303a_form.cleaned_data

                    # -------------------
                    update_pf_and_bill(rmn, cd)
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 303A Payment')

            if rmn == '304A':
                rm304a_form = PFormRM304A(data=request.POST, prefix='rm304a')
                if rm304a_form.is_valid():
                    cd = rm304a_form.cleaned_data

                    # -------------------
                    update_pf_and_bill(rmn, cd)
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 304A Payment')

            if rmn == '305A':
                rm305a_form = PFormRM305A(data=request.POST, prefix='rm305a')
                if rm305a_form.is_valid():
                    cd = rm305a_form.cleaned_data

                    # -------------------
                    update_pf_and_bill(rmn, cd)
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 305A Payment')

            if rmn == '306A':
                rm306a_form = PFormRM306A(data=request.POST, prefix='rm306a')
                if rm306a_form.is_valid():
                    cd = rm306a_form.cleaned_data

                    # -------------------
                    update_pf_and_bill(rmn, cd)
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 306A Payment')

            if rmn == '201B':
                rm201b_form = PFormRM201B(data=request.POST, prefix='rm201b')
                if rm201b_form.is_valid():
                    cd = rm201b_form.cleaned_data

                    # -------------------
                    update_pf_and_bill(rmn, cd)
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 201B Payment')

            if rmn == '202B':
                rm202b_form = PFormRM202B(data=request.POST, prefix='rm202b')
                if rm202b_form.is_valid():
                    cd = rm202b_form.cleaned_data

                    # -------------------
                    update_pf_and_bill(rmn, cd)
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 202B Payment')

            if rmn == '203B':
                rm203b_form = PFormRM203B(data=request.POST, prefix='rm203b')
                if rm203b_form.is_valid():
                    cd = rm203b_form.cleaned_data

                    # -------------------
                    update_pf_and_bill(rmn, cd)
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 203B Payment')

            if rmn == '204B':
                rm204b_form = PFormRM204B(data=request.POST, prefix='rm204b')
                if rm204b_form.is_valid():
                    cd = rm204b_form.cleaned_data

                    # -------------------
                    update_pf_and_bill(rmn, cd)
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 204B Payment')

            if rmn == '205B':
                rm205b_form = PFormRM205B(data=request.POST, prefix='rm205b')
                if rm205b_form.is_valid():
                    cd = rm205b_form.cleaned_data

                    # -------------------
                    update_pf_and_bill(rmn, cd)
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 205B Payment')

            if rmn == '301B':
                rm301b_form = PFormRM301B(data=request.POST, prefix='rm301b')
                if rm301b_form.is_valid():
                    cd = rm301b_form.cleaned_data

                    # -------------------
                    update_pf_and_bill(rmn, cd)
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 301B Payment')

            if rmn == '302B':
                rm302b_form = PFormRM302B(data=request.POST, prefix='rm302b')
                if rm302b_form.is_valid():
                    cd = rm302b_form.cleaned_data

                    # -------------------
                    update_pf_and_bill(rmn, cd)
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 302B Payment')

            if rmn == '303B':
                rm303b_form = PFormRM303B(data=request.POST, prefix='rm303b')
                if rm303b_form.is_valid():
                    cd = rm303b_form.cleaned_data

                    # -------------------
                    update_pf_and_bill(rmn, cd)
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 303B Payment')

            if rmn == '304B':
                rm304b_form = PFormRM304B(data=request.POST, prefix='rm304b')
                if rm304b_form.is_valid():
                    cd = rm304b_form.cleaned_data

                    # -------------------
                    update_pf_and_bill(rmn, cd)
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 304B Payment')

            if rmn == '305B':
                rm305b_form = PFormRM305B(data=request.POST, prefix='rm305b')
                if rm305b_form.is_valid():
                    cd = rm305b_form.cleaned_data

                    # -------------------
                    update_pf_and_bill(rmn, cd)
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 305B Payment')

            if rmn == '401B':
                rm401b_form = PFormRM401B(data=request.POST, prefix='rm401b')
                if rm401b_form.is_valid():
                    cd = rm401b_form.cleaned_data

                    # -------------------
                    update_pf_and_bill(rmn, cd)
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 401B Payment')

            if rmn == '402B':
                rm402b_form = PFormRM402B(data=request.POST, prefix='rm402b')
                if rm402b_form.is_valid():
                    cd = rm402b_form.cleaned_data

                    # -------------------
                    update_pf_and_bill(rmn, cd)
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 402B Payment')

            if rmn == '403B':
                rm403b_form = PFormRM403B(data=request.POST, prefix='rm403b')
                if rm403b_form.is_valid():
                    cd = rm403b_form.cleaned_data

                    # -------------------
                    update_pf_and_bill(rmn, cd)
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 403B Payment')

            if rmn == '404B':
                rm404b_form = PFormRM404B(data=request.POST, prefix='rm404b')
                if rm404b_form.is_valid():
                    cd = rm404b_form.cleaned_data

                    # -------------------
                    update_pf_and_bill(rmn, cd)
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 404B Payment')

            if rmn == '405B':
                rm405b_form = PFormRM405B(data=request.POST, prefix='rm405b')
                if rm405b_form.is_valid():
                    cd = rm405b_form.cleaned_data

                    # -------------------
                    update_pf_and_bill(rmn, cd)
                    # ------------------
                else:
                    messages.error(request, 'Error updating Room 405B Payment')


        else:
            if rmn == '101A':
                rm101a_form = PFormRM101A(prefix='rm101a')

            if rmn == '102A':
                rm102a_form = PFormRM102A(prefix='rm102a')

            if rmn == '103A':
                rm103a_form = PFormRM103A(prefix='rm103a')

            if rmn == '104A':
                rm104a_form = PFormRM104A(prefix='rm104a')

            if rmn == '105A':
                rm105a_form = PFormRM105A(prefix='rm105a')

            if rmn == '106A':
                rm106a_form = PFormRM106A(prefix='rm106a')

            if rmn == '201A':
                rm201a_form = PFormRM201A(prefix='rm201a')

            if rmn == '202A':
                rm202a_form = PFormRM202A(prefix='rm202a')

            if rmn == '203A':
                rm203a_form = PFormRM203A(prefix='rm203a')

            if rmn == '204A':
                rm204a_form = PFormRM204A(prefix='rm204a')

            if rmn == '205A':
                rm205a_form = PFormRM205A(prefix='rm205a')

            if rmn == '206A':
                rm206a_form = PFormRM206A(prefix='rm206a')

            if rmn == '301A':
                rm301a_form = PFormRM301A(prefix='rm301a')

            if rmn == '302A':
                rm302a_form = PFormRM302A(prefix='rm302a')

            if rmn == '303A':
                rm303a_form = PFormRM303A(prefix='rm303a')

            if rmn == '304A':
                rm304a_form = PFormRM304A(prefix='rm304a')

            if rmn == '305A':
                rm305a_form = PFormRM305A(prefix='rm305a')

            if rmn == '306A':
                rm306a_form = PFormRM306A(prefix='rm306a')

            if rmn == '201B':
                rm201b_form = PFormRM201B(prefix='rm201b')

            if rmn == '202B':
                rm202b_form = PFormRM202B(prefix='rm202b')

            if rmn == '203B':
                rm203b_form = PFormRM203B(prefix='rm203b')

            if rmn == '204B':
                rm204b_form = PFormRM204B(prefix='rm204b')

            if rmn == '205B':
                rm205b_form = PFormRM205B(prefix='rm205b')

            if rmn == '301B':
                rm301b_form = PFormRM301B(prefix='rm301b')

            if rmn == '302B':
                rm302b_form = PFormRM302B(prefix='rm302b')

            if rmn == '303B':
                rm303b_form = PFormRM303B(prefix='rm303b')

            if rmn == '304B':
                rm304b_form = PFormRM304B(prefix='rm304b')

            if rmn == '305B':
                rm305b_form = PFormRM305B(prefix='rm305b')

            if rmn == '401B':
                rm401b_form = PFormRM401B(prefix='rm401b')

            if rmn == '402B':
                rm402b_form = PFormRM402B(prefix='rm402b')

            if rmn == '403B':
                rm403b_form = PFormRM403B(prefix='rm403b')

            if rmn == '404B':
                rm404b_form = PFormRM404B(prefix='rm404b')

            if rmn == '405B':
                rm405b_form = PFormRM405B(prefix='rm405b')

    if request.method == 'POST':
        # messages.info(request, 'Payments have been completed !!!')
        messages.success(request, 'All payments have been completed !!!')

        return HttpResponseRedirect(reverse_lazy('admin_page'))
    else:
        return render(request, 'ams/payment.html', {'bills': bills, 'section': 'payment',
                                                    'rm101a_form': rm101a_form,
                                                    'rm102a_form': rm102a_form,
                                                    'rm103a_form': rm103a_form,
                                                    'rm104a_form': rm104a_form,
                                                    'rm105a_form': rm105a_form,
                                                    'rm106a_form': rm106a_form,

                                                    'rm201a_form': rm201a_form,
                                                    'rm202a_form': rm202a_form,
                                                    'rm203a_form': rm203a_form,
                                                    'rm204a_form': rm204a_form,
                                                    'rm205a_form': rm205a_form,
                                                    'rm206a_form': rm206a_form,

                                                    'rm301a_form': rm301a_form,
                                                    'rm302a_form': rm302a_form,
                                                    'rm303a_form': rm303a_form,
                                                    'rm304a_form': rm304a_form,
                                                    'rm305a_form': rm305a_form,
                                                    'rm306a_form': rm306a_form,

                                                    'rm201b_form': rm201b_form,
                                                    'rm202b_form': rm202b_form,
                                                    'rm203b_form': rm203b_form,
                                                    'rm204b_form': rm204b_form,
                                                    'rm205b_form': rm205b_form,

                                                    'rm301b_form': rm301b_form,
                                                    'rm302b_form': rm302b_form,
                                                    'rm303b_form': rm303b_form,
                                                    'rm304b_form': rm304b_form,
                                                    'rm305b_form': rm305b_form,

                                                    'rm401b_form': rm401b_form,
                                                    'rm402b_form': rm402b_form,
                                                    'rm403b_form': rm403b_form,
                                                    'rm404b_form': rm404b_form,
                                                    'rm405b_form': rm405b_form,

                                                    })


@login_required
def report_type(request):
    return render(request, 'ams/report_type.html', {'section': 'report'})


@login_required
def report_parameters(request):
    return render(request, 'ams/report_parameters.html', {'section': 'report'})


def get_eng_month_name(m: int):
    md = {1: 'January', 2: 'February', 3: 'March', 4: 'April', 5: 'May', 6: 'June', 7: 'July', 8: 'August',
          9: 'September',
          10: 'October', 11: 'November', 12: 'December'}
    im = int(m)
    return md[im]


def get_thai_month_name(bill_date: str):
    md = {1: 'มกราคม', 2: 'กุมภาพันธ์', 3: 'มีนาคม', 4: 'เมษายน', 5: 'พฤษภาคม', 6: 'มิถุนายน', 7: 'กรกฏาคม',
          8: 'สิงหาคม', 9: 'กันยายน',
          10: 'ตุลาคม', 11: 'พฤศจิกายน', 12: 'ธันวาคม'}

    y, m, d = bill_date.split('-')

    im = int(m)
    return md[im]


def get_thai_year(bill_date: str):
    y, m, d = bill_date.split('-')

    christ_y = int(y)
    buddist_y = christ_y + 543

    return str(buddist_y)


def get_aware_datetime(date_str):
    ret = parse_datetime(date_str)
    if not is_aware(ret):
        ret = make_aware(ret)
    return ret


# FULL MONTHLY REPORT ------------------------------------------------------------------
@login_required
def monthly_report(request):
    bld = request.POST['bld']
    if bld == 'AB':
        bld = 'A&B'

    mnth = int(request.POST['month'])
    mnth_name = get_eng_month_name(mnth)
    yr = int(request.POST['year'])

    # start_date = datetime.date(yr, mnth, 1)
    # end_date = datetime.date(yr, mnth, 30) # USE 30 TO AVOID OUT-OF-IDX RANGE

    # --------------------------------------------------------------------------
    start_date = datetime.datetime(yr, mnth, 1)
    end_date = datetime.datetime(yr, mnth, 30)  # USE 30 TO AVOID OUT OF INDX RANGE

    start_date = start_date.date()
    end_date = end_date.date()
    # --------------------------------------------------------------------------

    opl_a = None
    opl_b = None
    if bld == 'A':
        opl_a = Billing.objects.filter(status='close', room_no__endswith='A',
                                       bill_date__range=(start_date, end_date)).order_by('room_no')

    if bld == 'B':
        opl_b = Billing.objects.filter(status='close', room_no__endswith='B',
                                       bill_date__range=(start_date, end_date)).order_by('room_no')

    if bld == 'A&B':
        opl_a = Billing.objects.filter(status='close', room_no__endswith='A',
                                       bill_date__range=(start_date, end_date)).order_by('room_no')

        opl_b = Billing.objects.filter(status='close', room_no__endswith='B',

                                       bill_date__range=(start_date, end_date)).order_by('room_no')

    cum_list_a = []
    cum_list_b = []
    cum_list_ab = []

    trc_a = 0
    trac_a = 0
    tec_a = 0
    twc_a = 0
    tcsc_a = 0
    tosc_a = 0
    tovd_a = 0

    tlf_a = 0
    tma_a = 0

    tadj_a = 0

    tbt_a = 0
    tpa_a = 0
    tbf_a = 0

    trc_b = 0
    trac_b = 0
    tec_b = 0
    twc_b = 0
    tcsc_b = 0
    tosc_b = 0
    tovd_b = 0

    tlf_b = 0
    tma_b = 0

    tadj_b = 0

    tbt_b = 0
    tpa_b = 0
    tbf_b = 0

    trc_ab = 0
    trac_ab = 0
    tec_ab = 0
    twc_ab = 0
    tcsc_ab = 0
    tosc_ab = 0
    tovd_ab = 0

    tlf_ab = 0
    tma_ab = 0

    tadj_ab = 0

    tbt_ab = 0
    tpa_ab = 0
    tbf_ab = 0

    if opl_a:
        for bill in opl_a:
            trc_a += bill.room_cost
            trac_a += bill.room_acc_cost
            tec_a += bill.electricity_cost
            twc_a += bill.water_cost
            tcsc_a += bill.common_ser_cost
            tosc_a += bill.other_ser_cost
            tovd_a += bill.overdue_amount

            tlf_a += bill.late_fee
            tma_a += bill.maint_cost

            tadj_a += bill.adjust

            tbt_a += bill.bill_total
            tpa_a += bill.payment_amount
            tbf_a += bill.cf_amount
        cum_list_a.append(trc_a)
        cum_list_a.append(trac_a)
        cum_list_a.append(tec_a)
        cum_list_a.append(twc_a)
        cum_list_a.append(tcsc_a)
        cum_list_a.append(tosc_a)
        cum_list_a.append(tovd_a)

        cum_list_a.append(tlf_a)
        cum_list_a.append(tma_a)

        cum_list_a.append(tadj_a)

        cum_list_a.append(tbt_a)
        cum_list_a.append(tpa_a)
        cum_list_a.append(tbf_a)

    if opl_b:
        for bill in opl_b:
            trc_b += bill.room_cost
            trac_b += bill.room_acc_cost
            tec_b += bill.electricity_cost
            twc_b += bill.water_cost
            tcsc_b += bill.common_ser_cost
            tosc_b += bill.other_ser_cost
            tovd_b += bill.overdue_amount

            tlf_b += bill.late_fee
            tma_b += bill.maint_cost

            tadj_b += bill.adjust

            tbt_b += bill.bill_total
            tpa_b += bill.payment_amount
            tbf_b += bill.cf_amount
        cum_list_b.append(trc_b)
        cum_list_b.append(trac_b)
        cum_list_b.append(tec_b)
        cum_list_b.append(twc_b)
        cum_list_b.append(tcsc_b)
        cum_list_b.append(tosc_b)
        cum_list_b.append(tovd_b)

        cum_list_b.append(tlf_b)
        cum_list_b.append(tma_b)

        cum_list_b.append(tadj_b)

        cum_list_b.append(tbt_b)
        cum_list_b.append(tpa_b)
        cum_list_b.append(tbf_b)

    if opl_a and opl_b:
        trc_ab = trc_a + trc_b
        trac_ab = trac_a + trac_b

        tec_ab = tec_a + tec_b
        twc_ab = twc_a + twc_b
        tcsc_ab = tcsc_a + tcsc_b
        tosc_ab = tosc_a + tosc_b
        tovd_ab = tovd_a + tovd_b

        tlf_ab = tlf_a + tlf_b
        tma_ab = tma_a + tma_b

        tadj_ab = tadj_a + tadj_b

        tbt_ab = tbt_a + tbt_b
        tpa_ab = tpa_a + tpa_b
        tbf_ab = tbf_a + tbf_b

    return render(request, 'ams/monthly_report.html', {'opl_a': opl_a,
                                                       'opl_b': opl_b,
                                                       'bld': bld,
                                                       'mnth_name': mnth_name,
                                                       'yr': yr,

                                                       'cum_list_a': cum_list_a,
                                                       'cum_list_b': cum_list_b,
                                                       'cum_list_ab': cum_list_ab,
                                                       'trc_a': trc_a,
                                                       'trac_a': trac_a,
                                                       'tec_a': tec_a,
                                                       'twc_a': twc_a,
                                                       'tcsc_a': tcsc_a,
                                                       'tosc_a': tosc_a,
                                                       'tovd_a': tovd_a,

                                                       'tlf_a': tlf_a,
                                                       'tma_a': tma_a,

                                                       'tadj_a': tadj_a,

                                                       'tbt_a': tbt_a,
                                                       'tpa_a': tpa_a,
                                                       'tbf_a': tbf_a,

                                                       'trc_b': trc_b,
                                                       'trac_b': trac_b,
                                                       'tec_b': tec_b,
                                                       'twc_b': twc_b,
                                                       'tcsc_b': tcsc_b,
                                                       'tosc_b': tosc_b,
                                                       'tovd_b': tovd_b,

                                                       'tlf_b': tlf_b,
                                                       'tma_b': tma_b,

                                                       'tadj_b': tadj_b,

                                                       'tbt_b': tbt_b,
                                                       'tpa_b': tpa_b,
                                                       'tbf_b': tbf_b,

                                                       'trc_ab': trc_ab,
                                                       'trac_ab': trac_ab,
                                                       'tec_ab': tec_ab,
                                                       'twc_ab': twc_ab,
                                                       'tcsc_ab': tcsc_ab,
                                                       'tosc_ab': tosc_ab,
                                                       'tovd_ab': tovd_ab,

                                                       'tlf_ab': tlf_ab,
                                                       'tma_ab': tma_ab,

                                                       'tadj_ab': tadj_ab,

                                                       'tbt_ab': tbt_ab,
                                                       'tpa_ab': tpa_ab,
                                                       'tbf_ab': tbf_ab,

                                                       })


# MINI MONTHLY REPORT ------------------------------------------------------------------

@login_required
def monthly_report_mini(request):
    bld = request.POST['bld']
    if bld == 'AB':
        bld = 'A&B'

    mnth = int(request.POST['month'])
    mnth_name = get_eng_month_name(mnth)
    yr = int(request.POST['year'])

    # start_date = datetime.date(yr, mnth, 1)
    # end_date = datetime.date(yr, mnth, 30) # USE 30 TO AVOID OUT-OF-IDX RANGE

    # --------------------------------------------------------------------------
    start_date = datetime.datetime(yr, mnth, 1)
    end_date = datetime.datetime(yr, mnth, 30)  # USE 30 TO AVOID OUT OF INDX RANGE

    start_date = start_date.date()
    end_date = end_date.date()
    # --------------------------------------------------------------------------

    opl_a = None
    opl_b = None
    if bld == 'A':
        opl_a = Billing.objects.filter(status='close', room_no__endswith='A',
                                       bill_date__range=(start_date, end_date)).order_by('room_no')

    if bld == 'B':
        opl_b = Billing.objects.filter(status='close', room_no__endswith='B',
                                       bill_date__range=(start_date, end_date)).order_by('room_no')

    if bld == 'A&B':
        opl_a = Billing.objects.filter(status='close', room_no__endswith='A',
                                       bill_date__range=(start_date, end_date)).order_by('room_no')

        opl_b = Billing.objects.filter(status='close', room_no__endswith='B',

                                       bill_date__range=(start_date, end_date)).order_by('room_no')

    trcac_a = 0

    tec_a = 0
    twc_a = 0
    tcsc_a = 0
    tosc_a = 0
    tovd_a = 0

    tlf_ma_a = 0

    tbt_a = 0
    tpa_a = 0

    trcac_b = 0

    tec_b = 0
    twc_b = 0
    tcsc_b = 0
    tosc_b = 0
    tovd_b = 0

    tlf_ma_b = 0

    tbt_b = 0
    tpa_b = 0

    trcac_ab = 0

    tec_ab = 0
    twc_ab = 0
    tcsc_ab = 0
    tosc_ab = 0
    tovd_ab = 0

    tlf_ma_ab = 0

    tbt_ab = 0
    tpa_ab = 0

    if opl_a:
        for bill in opl_a:
            trcac_a += (bill.room_cost + bill.room_acc_cost + bill.adjust)

            tec_a += bill.electricity_cost
            twc_a += bill.water_cost
            tcsc_a += bill.common_ser_cost
            tosc_a += bill.other_ser_cost
            tovd_a += bill.overdue_amount

            tlf_ma_a += (bill.late_fee + bill.maint_cost)

            tbt_a += bill.bill_total
            tpa_a += bill.payment_amount

    if opl_b:
        for bill in opl_b:
            trcac_b += (bill.room_cost + bill.room_acc_cost + bill.adjust)

            tec_b += bill.electricity_cost
            twc_b += bill.water_cost
            tcsc_b += bill.common_ser_cost
            tosc_b += bill.other_ser_cost
            tovd_b += bill.overdue_amount

            tlf_ma_b += (bill.late_fee + bill.maint_cost)

            tbt_b += bill.bill_total
            tpa_b += bill.payment_amount

    if opl_a and opl_b:
        trcac_ab = trcac_a + trcac_b

        tec_ab = tec_a + tec_b
        twc_ab = twc_a + twc_b
        tcsc_ab = tcsc_a + tcsc_b
        tosc_ab = tosc_a + tosc_b
        tovd_ab = tovd_a + tovd_b

        tlf_ma_ab = tlf_ma_a + tlf_ma_b

        tbt_ab = tbt_a + tbt_b
        tpa_ab = tpa_a + tpa_b

    return render(request, 'ams/monthly_report_mini.html', {'opl_a': opl_a,
                                                            'opl_b': opl_b,
                                                            'bld': bld,
                                                            'mnth_name': mnth_name,
                                                            'yr': yr,

                                                            'trcac_a': trcac_a,

                                                            'tec_a': tec_a,
                                                            'twc_a': twc_a,
                                                            'tcsc_a': tcsc_a,
                                                            'tosc_a': tosc_a,
                                                            'tovd_a': tovd_a,

                                                            'tlf_ma_a': tlf_ma_a,

                                                            'tbt_a': tbt_a,
                                                            'tpa_a': tpa_a,

                                                            'trcac_b': trcac_b,

                                                            'tec_b': tec_b,
                                                            'twc_b': twc_b,
                                                            'tcsc_b': tcsc_b,
                                                            'tosc_b': tosc_b,
                                                            'tovd_b': tovd_b,

                                                            'tlf_ma_b': tlf_ma_b,

                                                            'tbt_b': tbt_b,
                                                            'tpa_b': tpa_b,

                                                            'trcac_ab': trcac_ab,

                                                            'tec_ab': tec_ab,
                                                            'twc_ab': twc_ab,
                                                            'tcsc_ab': tcsc_ab,
                                                            'tosc_ab': tosc_ab,
                                                            'tovd_ab': tovd_ab,

                                                            'tlf_ma_ab': tlf_ma_ab,

                                                            'tbt_ab': tbt_ab,
                                                            'tpa_ab': tpa_ab,

                                                            })


# -------------------------------------------------------------------------------------------

@login_required
def extra_service(request):
    extra = Extra.objects.all().order_by('id')

    current_dt = datetime.datetime.now()

    return render(request, 'ams/extra_service.html', {'extra': extra, 'current_dt': current_dt})


@login_required
def elec_cpu_change(request):
    if request.method == 'POST':
        elec_cpu_form = Elec_cpu_change(request.POST)
        if elec_cpu_form.is_valid():
            cd = elec_cpu_form.cleaned_data

            ex_item = get_object_or_404(Extra, desc='Electricity CPU')
            ex_item.cpu = cd['elec_cpu']
            ex_item.save()

            messages.info(request, 'Electricity CPU has been chnaged !!')

            return HttpResponseRedirect(reverse_lazy('admin_page'))
        else:
            messages.ERROR(request, 'Error ... !!')
    else:
        elec_cpu_form = Elec_cpu_change()
    return render(request, 'ams/elec_cpu_change.html', {'elec_cpu_form': elec_cpu_form})


@login_required
def water_cpu_change(request):
    if request.method == 'POST':
        water_cpu_form = Water_cpu_change(request.POST)
        if water_cpu_form.is_valid():
            cd = water_cpu_form.cleaned_data

            ex_item = get_object_or_404(Extra, desc='Water CPU')
            ex_item.cpu = cd['water_cpu']
            ex_item.save()

            messages.success(request, 'Water CPU has been chnaged !!')
            return HttpResponseRedirect(reverse_lazy('admin_page'))
        else:
            messages.ERROR(request, 'Error ... !!')
    else:
        water_cpu_form = Water_cpu_change()
    return render(request, 'ams/water_cpu_change.html', {'water_cpu_form': water_cpu_form})


@login_required
def room_type_rate(request):
    rm_type_rate = Room_type.objects.all()
    current_dt = datetime.datetime.now()

    return render(request, 'ams/room_type_rate.html', {'rm_type_rate': rm_type_rate, 'current_dt': current_dt})


@login_required
def current_tenant(request):
    cur_tenant = TenantProfile.objects.all().order_by('start_date')

    total_tn = cur_tenant.count()

    current_dt = datetime.datetime.now()

    return render(request, 'ams/current_tenant.html',
                  {'cur_tenant': cur_tenant, 'current_dt': current_dt, 'total_tn': total_tn})


@login_required
def vacant_rooms(request):
    current_dt = datetime.datetime.now()

    all_room = Room.objects.all()
    cur_tn = TenantProfile.objects.all()
    oc_rm_set = []
    vac_rm_set = []
    for tn in cur_tn:
        oc_rm_set.append(tn.room_no.room_no)

    for rm in all_room:
        if rm.room_no not in oc_rm_set:
            vac_rm_set.append(rm.room_no)

    return render(request, 'ams/vacant_rooms.html', {'vac_rm_set': vac_rm_set, 'current_dt': current_dt})


# --------------------------------------------

@login_required
def send_sms_confirmation(request):
    open_bill = Billing.objects.filter(status='open').order_by('id')
    active_tenant = TenantProfile.objects.all()

    total_opb = open_bill.count()
    total_tenant = active_tenant.count()

    return render(request, 'ams/confirmation.html', {'total_opb': total_opb, 'total_tenant': total_tenant})


@login_required
def send_sms_execution(request):
    val = request.POST['radbt']

    if val == 'ng':
        return render(request, 'ams/misc_contents.html')
    else:
        send_bill_sms_to_all_tenants(request)
        return HttpResponseRedirect(reverse_lazy('misc_contents'))


# --------------------------------------------


# SENDING MESSAGE FROM LOCALHOST AND FROM WEB !!!! **********************************************************
def send_message(to_phone_no, msg):
    account_sid = GV.Account_SID
    auth_token = GV.Auth_Token
    client = Client(account_sid, auth_token)
    sending_phone_no = GV.Sending_Phone_No

    tenant_phone_no = '+66' + to_phone_no
    sending_message = msg

    # SENDING MESSAGE *********************************************************************************
    message = client.messages.create(to=tenant_phone_no, from_=sending_phone_no, body=sending_message)

    # *************************************************************************************************


# # SENDING MESSAGE FROM WEB ??? ****************************************************************************
# def send_message(to_phone_no, msg):

#     proxy_client = TwilioHttpClient()
#     proxy_client.session.proxies = {'https':os.environ['https_proxy']}

#     account_sid = GV.Account_SID
#     auth_token = GV.Auth_Token

#     client = Client(account_sid, auth_token, http_client=proxy_client)

#     sending_phone_no = GV.Sending_Phone_No
#     tenant_phone_no = '+66' + to_phone_no
#     sending_message = msg

#     # SENDING MESSAGE ********************************************************************************
#     message = client.messages.create(to=tenant_phone_no, from_=sending_phone_no, body=sending_message)
#     # ************************************************************************************************


@login_required
def send_bill_sms_to_all_tenants(request):
    bills = Billing.objects.filter(status='open').order_by('id')
    total_tenant = TenantProfile.objects.all().count()

    if bills:

        total_open_bills = total_tenant
        no_of_bills_sent = 0
        for rmn_bill in bills:
            rmn = rmn_bill.room_no

            rmn_pf = get_object_or_404(TenantProfile, room_no__room_no=rmn)

            # --------------------------------------------------
            if GV.SMS_TO_ALL_ROOMS:
                rmn_hp = rmn_pf.phone
            else:
                rmn_hp = '0840860087'  # TESTING ONLY'
            # -------------------------------------------------

            bill_dt = rmn_bill.bill_date
            cur_mth = bill_dt.month
            cur_yr = bill_dt.year
            cur_th_mth = get_thai_month_name(str(bill_dt))
            cur_th_yr = get_thai_year(str(bill_dt))
            next_dt_mth = datetime.date(cur_yr, cur_mth + 1, 15)
            next_th_m = get_thai_month_name(str(next_dt_mth))

            rn = rmn
            bank_info = GV.bank_info

            rc_and_rac = rmn_bill.room_cost + rmn_bill.room_acc_cost + rmn_bill.adjust
            wc = rmn_bill.water_cost
            ec = rmn_bill.electricity_cost
            csc = rmn_bill.common_ser_cost
            osc = rmn_bill.other_ser_cost
            ovd = rmn_bill.overdue_amount

            # TEMPORARY UNTIL OVD OF RM204A HAS BEEN COVERED =========
            if rmn == '204A':
                bt = rmn_bill.bill_total - ovd + 1000
            else:
                bt = rmn_bill.bill_total
            # ========================================================

            bill_msg = "<ห้อง {0}> กรุณาชำระค่าเช่า ประจำเดือน {1} จำนวน {2:,.0f} บาท ภายในวันที่ 5 {3} {4} {5} <รายละเอียดค่าเช่า> ห้องพร้อมอุปกรณ์: {6:,.0f} บาท นำ้: {7:,.0f} บาท ไฟ: {8:,.0f} บาท ส่วนกลาง: {9:,.0f} บาท อื่นๆ: {10:,.0f} บาท ค้างจ่าย: {11:,.0f} บาท"
            bill_msg = bill_msg.format(rn, cur_th_mth, bt, next_th_m, cur_th_yr, bank_info, rc_and_rac, wc, ec, csc,
                                       osc,
                                       ovd)

            # print(rmn_hp, ': ', bill_msg)

            # -------------------------------
            send_message(rmn_hp, bill_msg)
            # -------------------------------
            no_of_bills_sent += 1

        messages.success(request,
                         'Billing SMS has been sent to: {} of {} rooms !!'.format(no_of_bills_sent, total_open_bills))

    else:
        messages.info(request, 'No open bills are available !!!')

    return HttpResponseRedirect(reverse_lazy('misc_contents'))


@login_required
def send_sms_to_individual_room(request):
    bills = Billing.objects.filter(status='open').order_by('id')

    if bills:
        if request.method == 'POST':
            rn = request.POST['rmn']
            rmn_bill = get_object_or_404(Billing, room_no=rn, status='open')

            bill_dt = rmn_bill.bill_date
            cur_mth = bill_dt.month
            cur_yr = bill_dt.year
            cur_th_mth = get_thai_month_name(str(bill_dt))
            cur_th_yr = get_thai_year(str(bill_dt))
            next_dt_mth = datetime.date(cur_yr, cur_mth + 1, 15)
            next_th_m = get_thai_month_name(str(next_dt_mth))
            bank_info = GV.bank_info

            rc_and_rac = rmn_bill.room_cost + rmn_bill.room_acc_cost + rmn_bill.adjust
            wc = rmn_bill.water_cost
            ec = rmn_bill.electricity_cost
            csc = rmn_bill.common_ser_cost
            osc = rmn_bill.other_ser_cost
            ovd = rmn_bill.overdue_amount

            # TEMPORARY UNTIL OVD OF RM204A HAS BEEN COVERED
            if rn == '204A':
                bt = rmn_bill.bill_total - ovd + 1000
            else:
                bt = rmn_bill.bill_total
            # ===============================================

            rmn_pf = get_object_or_404(TenantProfile, room_no__room_no=rn)

            # to_phone_no = rmn_pf.phone # To be permanent
            to_phone_no = '0840860087'  # TESTING ONLY

            bill_msg = "<ห้อง {0}> กรุณาชำระค่าเช่า ประจำเดือน {1} จำนวน {2:,.0f} บาท ภายในวันที่ 5 {3} {4} {5} <รายละเอียดค่าเช่า> ห้องพร้อมอุปกรณ์: {6:,.0f} บาท นำ้: {7:,.0f} บาท ไฟ: {8:,.0f} บาท ส่วนกลาง: {9:,.0f} บาท อื่นๆ: {10:,.0f} บาท ค้างจ่าย: {11:,.0f} บาท"
            bill_msg = bill_msg.format(rn, cur_th_mth, bt, next_th_m, cur_th_yr, bank_info, rc_and_rac, wc, ec, csc,
                                       osc,
                                       ovd)

            # --------------------------------------
            send_message(to_phone_no, bill_msg)
            # --------------------------------------

            messages.success(request, 'Billing SMS has been sent to: {}-{} !!'.format(rn, to_phone_no))
            return HttpResponseRedirect(reverse_lazy('misc_contents'))
        else:
            return render(request, 'ams/send_sms_to_individual_room.html', {'bills': bills})
    else:
        messages.info(request, 'No open bills are available !!!')
    return HttpResponseRedirect(reverse_lazy('misc_contents'))


@login_required
def send_general_sms(request):
    if request.method == 'POST':
        phone_msg_form = PhoneNoMessage(request.POST)
        if phone_msg_form.is_valid():
            cd = phone_msg_form.cleaned_data

            phn = cd['phone_no']
            msg = cd['sms_msg']

            try:
                send_message(phn, msg)
            except Exception as err:
                messages.error(request, 'ERROR: {}'.format(str(err)))
                return HttpResponseRedirect(reverse_lazy('misc_contents'))
            else:
                messages.success(request, 'SMS has been sent to: {} !!'.format(phn))
                return HttpResponseRedirect(reverse_lazy('misc_contents'))
    else:
        phone_msg_form = PhoneNoMessage()
        return render(request, 'ams/send_general_sms.html', {'phone_msg_form': phone_msg_form})


@login_required
def misc_contents(request):
    return render(request, 'ams/misc_contents.html', {'section': 'misc'})


@login_required
def tenant_page(request):
    usr = str(request.user)
    fn, ln = usr.split(" ")
    # tenant_pf = get_object_or_404(TenantProfile, tenant__first_name=fn, tenant__last_name=ln)
    try:
        tenant_pf = TenantProfile.objects.get(tenant__first_name=fn, tenant__last_name=ln)
    except Exception as err:
        messages.error(request, 'ERROR: {} '.format(str(err)))
        return HttpResponseRedirect(reverse_lazy('login'))
    else:
        exd = {}
        exd.setdefault('Electricity CPU', 0)
        exd.setdefault('Water CPU', 0)
        exd.setdefault('Garbage', 0)
        exd.setdefault('Parking', 0)
        exd.setdefault('Wifi', 0)
        exd.setdefault('Cable TV', 0)
        exd.setdefault('Bed', 0)
        exd.setdefault('Bed accessories', 0)
        exd.setdefault('Dressing Table', 0)
        exd.setdefault('Clothing Cupboard', 0)
        exd.setdefault('TV Table', 0)
        exd.setdefault('Fridge', 0)
        exd.setdefault('Air-Conditioner', 0)

        for e in tenant_pf.extra.all():
            exd.update({e.desc: e.cpu})

        room_acc_cost = exd['Bed'] + exd['Bed accessories'] + exd['Dressing Table'] \
                        + exd['Clothing Cupboard'] + exd['TV Table'] + exd['Fridge'] \
                        + exd['Air-Conditioner']

        oth_ser_cost = exd['Garbage'] + exd['Parking'] + exd['Wifi'] + exd['Cable TV']

        cur_dt = datetime.datetime.now()

        return render(request, 'ams/tenant_page.html',
                      {'section': 'tenant_profile', 'tenant_pf': tenant_pf, 'room_acc_cost': room_acc_cost,
                       'oth_ser_cost': oth_ser_cost, 'cur_dt': cur_dt})


def tenant_bill_subroutine(tn_bill):
    bill_dt = tn_bill.bill_date
    pay_date = tn_bill.payment_date
    cur_mth = bill_dt.month
    cur_yr = bill_dt.year
    cur_th_mth = get_thai_month_name(str(bill_dt))
    cur_th_yr = get_thai_year(str(bill_dt))
    next_dt_mth = datetime.date(cur_yr, cur_mth + 1, 15)
    next_th_m = get_thai_month_name(str(next_dt_mth))

    room_with_acc_cost = tn_bill.room_cost + tn_bill.room_acc_cost + tn_bill.adjust

    pay_amt = tn_bill.payment_amount

    bill_misc = tn_bill.late_fee + tn_bill.maint_cost

    if tn_bill.status == 'open':
        paid_str = 'รอชำระ'
    else:
        paid_str = 'ชำระแล้ว ณ วันที่ {0} {1} {2} จำนวน {3:,.0f} บาท'.format(pay_date.day,
                                                                             get_thai_month_name(str(pay_date)),
                                                                             get_thai_year(str(pay_date)), pay_amt)

    # TEMPORARY UNTIL OVD OF RM204A HAS BEEN COVERED
    rn = tn_bill.room_no
    if rn == '204A':
        bill_total = tn_bill.bill_total - tn_bill.overdue_amount + 1000
    else:
        bill_total = tn_bill.bill_total

    return room_with_acc_cost, bill_misc, bill_total, paid_str, cur_th_mth, next_th_m, cur_th_yr


@login_required
def tenant_bill(request):
    tenant = str(request.user)
    bills = Billing.objects.filter(tenant_name=tenant)

    if bills:
        tnb_qs = Billing.objects.filter(tenant_name=tenant, status='open')
        if tnb_qs:
            tn_bill = get_object_or_404(Billing, tenant_name=tenant, status='open')
        else:
            bill_month = str(datetime.datetime.now().month)
            tnb_qs = Billing.objects.filter(tenant_name=tenant, status='close', bill_date__month=bill_month)
            if tnb_qs:
                tn_bill = get_object_or_404(Billing, tenant_name=tenant, status='close', bill_date__month=bill_month)
            else:
                bill_month = str(datetime.datetime.now().month - 1)
                tn_bill = get_object_or_404(Billing, tenant_name=tenant, status='close', bill_date__month=bill_month)

        room_with_acc_cost, bill_misc, bill_total, paid_str, cur_th_mth, next_th_m, cur_th_yr = tenant_bill_subroutine(
            tn_bill)

        return render(request, 'ams/tenant_bill.html',
                      {'section': 'bill', 'tn_bill': tn_bill, 'room_with_acc_cost': room_with_acc_cost,
                       'bill_misc': bill_misc, 'bill_total': bill_total, 'cur_th_mth': cur_th_mth,
                       'next_th_m': next_th_m,
                       'cur_th_yr': cur_th_yr, 'paid_str': paid_str})
    else:

        # NEW TENANT
        return HttpResponseRedirect(reverse_lazy('new_tenant'))


# -------------------------------------------------------------------------------------------------


@login_required
def new_tenant(request):
    tenant_name = str(request.user)

    return render(request, 'ams/new_tenant.html', {'section': 'bill', 'tenant_name': tenant_name})


@login_required
def admin_page(request):
    return render(request, 'ams/admin_page.html')


def get_ref_string():
    char_str = list('ABCDEFGHIJKLMNOPQRSTUVWXYZ')
    random.shuffle(char_str)
    fd = random.choice(char_str)

    sd = str(random.randint(0, 9)) + str(random.randint(0, 9)) + str(random.randint(0, 9)) + str(random.randint(0, 9))
    ref_str = fd + '-' + sd

    return ref_str


def get_eng_month_name(m: int):
    md = {1: 'January', 2: 'February', 3: 'March', 4: 'April', 5: 'May', 6: 'June', 7: 'July', 8: 'August',
          9: 'September',
          10: 'October', 11: 'November', 12: 'December'}
    im = int(m)
    return md[im]


def get_thai_month_name(bill_date: str):
    md = {1: 'มกราคม', 2: 'กุมภาพันธ์', 3: 'มีนาคม', 4: 'เมษายน', 5: 'พฤษภาคม', 6: 'มิถุนายน', 7: 'กรกฏาคม',
          8: 'สิงหาคม', 9: 'กันยายน',
          10: 'ตุลาคม', 11: 'พฤศจิกายน', 12: 'ธันวาคม'}

    y, m, d = bill_date.split('-')

    im = int(m)
    return md[im]


def get_thai_year(bill_date: str):
    y, m, d = bill_date.split('-')

    christ_y = int(y)
    buddist_y = christ_y + 543

    return str(buddist_y)


def make_date_string(self, ds: str):
    y, m, d = str(ds).split('-')
    return d + '-' + m + '-' + y


def give_error_message(error_msg):
    print(error_msg)


def give_info_message(error_msg):
    print(error_msg)


# @login_required
def write_to_bill_summary(request, excel_f, opbl):
    excel_cell_dict = {
        'A': {'101': ['G15', 'H15', 'I15', 'J15', 'K15', 'L15', 'M15', 'N15', 'O15', 'P15', 'Q15', 'R15'],
              '102': ['G16', 'H16', 'I16', 'J16', 'K16', 'L16', 'M16', 'N16', 'O16', 'P16', 'Q16', 'R16'],
              '103': ['G17', 'H17', 'I17', 'J17', 'K17', 'L17', 'M17', 'N17', 'O17', 'P17', 'Q17', 'R17'],
              '104': ['G18', 'H18', 'I18', 'J18', 'K18', 'L18', 'M18', 'N18', 'O18', 'P18', 'Q18', 'R18'],
              '105': ['G19', 'H19', 'I19', 'J19', 'K19', 'L19', 'M19', 'N19', 'O19', 'P19', 'Q19', 'R19'],
              '106': ['G20', 'H20', 'I20', 'J20', 'K20', 'L20', 'M20', 'N20', 'O20', 'P20', 'Q20', 'R20'],
              '201': ['G21', 'H21', 'I21', 'J21', 'K21', 'L21', 'M21', 'N21', 'O21', 'P21', 'Q21', 'R21'],
              '202': ['G22', 'H22', 'I22', 'J22', 'K22', 'L22', 'M22', 'N22', 'O22', 'P22', 'Q22', 'R22'],
              '203': ['G23', 'H23', 'I23', 'J23', 'K23', 'L23', 'M23', 'N23', 'O23', 'P23', 'Q23', 'R23'],
              '204': ['G24', 'H24', 'I24', 'J24', 'K24', 'L24', 'M24', 'N24', 'O24', 'P24', 'Q24', 'R24'],
              '205': ['G25', 'H25', 'I25', 'J25', 'K25', 'L25', 'M25', 'N25', 'O25', 'P25', 'Q25', 'R25'],
              '206': ['G26', 'H26', 'I26', 'J26', 'K26', 'L26', 'M26', 'N26', 'O26', 'P26', 'Q26', 'R26'],
              '301': ['G27', 'H27', 'I27', 'J27', 'K27', 'L27', 'M27', 'N27', 'O27', 'P27', 'Q27', 'R27'],
              '302': ['G28', 'H28', 'I28', 'J28', 'K28', 'L28', 'M28', 'N28', 'O28', 'P28', 'Q28', 'R28'],
              '303': ['G29', 'H29', 'I29', 'J29', 'K29', 'L29', 'M29', 'N29', 'O29', 'P29', 'Q29', 'R29'],
              '304': ['G30', 'H30', 'I30', 'J30', 'K30', 'L30', 'M30', 'N30', 'O30', 'P30', 'Q30', 'R30'],
              '305': ['G31', 'H31', 'I31', 'J31', 'K31', 'L31', 'M31', 'N31', 'O31', 'P31', 'Q31', 'R31'],
              '306': ['G32', 'H32', 'I32', 'J32', 'K32', 'L32', 'M32', 'N32', 'O32', 'P32', 'Q32', 'R32']

              },

        'B': {'101': ['G15', 'H15', 'I15', 'J15', 'K15', 'L15', 'M15', 'N15', 'O15', 'P15', 'Q15', 'R15'],
              '102': ['G16', 'H16', 'I16', 'J16', 'K16', 'L16', 'M16', 'N16', 'O16', 'P16', 'Q16', 'R16'],
              '103': ['G17', 'H17', 'I17', 'J17', 'K17', 'L17', 'M17', 'N17', 'O17', 'P17', 'Q17', 'R17'],
              '104': ['G18', 'H18', 'I18', 'J18', 'K18', 'L18', 'M18', 'N18', 'O18', 'P18', 'Q18', 'R18'],
              '105': ['G19', 'H19', 'I19', 'J19', 'K19', 'L19', 'M19', 'N19', 'O19', 'P19', 'Q19', 'R19'],
              '106': ['G20', 'H20', 'I20', 'J20', 'K20', 'L20', 'M20', 'N20', 'O20', 'P20', 'Q20', 'R20'],
              '201': ['G21', 'H21', 'I21', 'J21', 'K21', 'L21', 'M21', 'N21', 'O21', 'P21', 'Q21', 'R21'],
              '202': ['G22', 'H22', 'I22', 'J22', 'K22', 'L22', 'M22', 'N22', 'O22', 'P22', 'Q22', 'R22'],
              '203': ['G23', 'H23', 'I23', 'J23', 'K23', 'L23', 'M23', 'N23', 'O23', 'P23', 'Q23', 'R23'],
              '204': ['G24', 'H24', 'I24', 'J24', 'K24', 'L24', 'M24', 'N24', 'O24', 'P24', 'Q24', 'R24'],
              '205': ['G25', 'H25', 'I25', 'J25', 'K25', 'L25', 'M25', 'N25', 'O25', 'P25', 'Q25', 'R25'],
              '206': ['G26', 'H26', 'I26', 'J26', 'K26', 'L26', 'M26', 'N26', 'O26', 'P26', 'Q26', 'R26'],
              '301': ['G27', 'H27', 'I27', 'J27', 'K27', 'L27', 'M27', 'N27', 'O27', 'P27', 'Q27', 'R27'],
              '302': ['G28', 'H28', 'I28', 'J28', 'K28', 'L28', 'M28', 'N28', 'O28', 'P28', 'Q28', 'R28'],
              '303': ['G29', 'H29', 'I29', 'J29', 'K29', 'L29', 'M29', 'N29', 'O29', 'P29', 'Q29', 'R29'],
              '304': ['G30', 'H30', 'I30', 'J30', 'K30', 'L30', 'M30', 'N30', 'O30', 'P30', 'Q30', 'R30'],
              '305': ['G31', 'H31', 'I31', 'J31', 'K31', 'L31', 'M31', 'N31', 'O31', 'P31', 'Q31', 'R31'],
              '306': ['G32', 'H32', 'I32', 'J32', 'K32', 'L32', 'M32', 'N32', 'O32', 'P32', 'Q32', 'R32'],
              '401': ['G33', 'H33', 'I33', 'J33', 'K33', 'L33', 'M33', 'N33', 'O33', 'P33', 'Q33', 'R33'],
              '402': ['G34', 'H34', 'I34', 'J34', 'K34', 'L34', 'M34', 'N34', 'O34', 'P34', 'Q34', 'R34'],
              '403': ['G35', 'H35', 'I35', 'J35', 'K35', 'L35', 'M35', 'N35', 'O35', 'P35', 'Q35', 'R35'],
              '404': ['G36', 'H36', 'I36', 'J36', 'K36', 'L36', 'M36', 'N36', 'O36', 'P36', 'Q36', 'R36'],
              '405': ['G37', 'H37', 'I37', 'J37', 'K37', 'L37', 'M37', 'N37', 'O37', 'P37', 'Q37', 'R37'],
              '406': ['G38', 'H38', 'I38', 'J38', 'K38', 'L38', 'M38', 'N38', 'O38', 'P38', 'Q38', 'R38']

              }

    }

    # LOAD EXCEL WORKING FILE FROM DISK INTO PYTHON

    cwd = os.getcwd()

    wb = load_workbook(excel_f)

    open_bill_date = opbl[0].bill_date  # from any open bill

    month = open_bill_date.month
    day = open_bill_date.day
    year = open_bill_date.year
    thai_year = get_thai_year(str(open_bill_date))

    # CLEAR EXCEL WORKBOOK
    clear_excel_bill_summary_worksheet(wb, excel_cell_dict, excel_f)

    # bdn = ''
    # rmn = ''
    excel_tab = ''

    for e in opbl:

        rmn = e.room_no[0:3]
        bdn = e.room_no[3]

        if bdn == 'A':
            excel_tab = 'B1SUMMARY'
        elif bdn == 'B':
            excel_tab = 'B2SUMMARY'

        # WRITE TO DATE BLOCK
        wb[excel_tab]['D4'] = day
        wb[excel_tab]['D5'] = month
        wb[excel_tab]['D6'] = thai_year

        # LIST OF EXCEL ELEES
        cell = excel_cell_dict[bdn][rmn]

        wb[excel_tab][cell[0]] = e.bill_ref
        wb[excel_tab][cell[1]] = e.tenant_name
        wb[excel_tab][cell[2]] = e.room_no[0:3]
        wb[excel_tab][cell[3]] = e.room_no[3]
        wb[excel_tab][cell[4]] = e.room_cost
        wb[excel_tab][cell[5]] = e.room_acc_cost
        wb[excel_tab][cell[6]] = e.electricity_cost
        wb[excel_tab][cell[7]] = e.water_cost
        wb[excel_tab][cell[8]] = e.common_ser_cost
        wb[excel_tab][cell[9]] = e.other_ser_cost
        wb[excel_tab][cell[10]] = e.adjust
        wb[excel_tab][cell[11]] = e.overdue_amount

    try:

        # SAVE TO DISK FILE
        cwd = os.getcwd()

        wb.save(excel_f)

        total_processed_rooms = len(opbl)

        messages.info(request,
                      'Bill Summary: {0}\\{1} with Total {2} Rooms SAVED !'.format(cwd, excel_f, total_processed_rooms))

    except Exception as err:
        give_error_message('Error: '.format(err))


def clear_excel_bill_summary_worksheet(wb, worksheet_cell_dict, excel_file):
    # FULL OPEN-BILL-LIST (ALL ROOMS OCCUOIED)

    bd1_room = ['101A', '102A', '103A', '104A', '105A', '106A',
                '201A', '202A', '203A', '204A', '205A', '206A',
                '301A', '302A', '303A', '304A', '305A', '306A']

    bd2_room = ['101B', '102B', '103B', '104B', '105B', '106B',
                '201B', '202B', '203B', '204B', '205B', '206B',
                '301B', '302B', '303B', '304B', '305B', '306B',
                '401B', '402B', '403B', '404B', '405B', '406B']

    bd1_room.extend(bd2_room)  # COMBINE ALL ROOMS FROM BLD1 & 2 TO bd1_room

    excel_tab = ''

    for e in bd1_room:

        bd_no = e[3]
        rm_no = e[0:3]

        if bd_no == 'A':
            excel_tab = 'B1SUMMARY'
        elif bd_no == 'B':
            excel_tab = 'B2SUMMARY'

        number_of_cell = 12  # NUMBER OF CELL to be cleared FOR EACH ROOM (EXCLUDE 'TOTAL')
        cell = worksheet_cell_dict[bd_no][rm_no]  # List of exel cells

        # CLEAR EXCELL WORKSHEET

        # CLEAR DATE_BLOCK
        wb[excel_tab]['D4'] = ''  # wb[tab]['D4']
        wb[excel_tab]['D5'] = ''
        wb[excel_tab]['D6'] = ''

        for i in range(number_of_cell):
            # CLEAR EXCEL-WORKSHEET
            wb[excel_tab][cell[i]] = ''

    try:
        wb.save(excel_file)  # SAVE EXCEL WORKBOOK TO DISK FILE
    except Exception as err:
        give_error_message('Error: {}'.format(err))


def get_floor_no(room_no):  # rm_no = 103A

    rm_no = room_no[0:3]  # 103
    bld_no = room_no[3]  # A

    flr_1 = ['101', '102', '103', '104', '105', '106']
    flr_2 = ['201', '202', '203', '204', '205', '206']
    flr_3 = ['301', '302', '303', '304', '305', '306']
    flr_4 = ['401', '402', '403', '404', '405', '406']

    flr_no = ''

    if bld_no == 'A':
        if rm_no in flr_1:
            flr_no = 'SB1F1'
        elif rm_no in flr_2:
            flr_no = 'SB1F2'
        elif rm_no in flr_3:
            flr_no = 'SB1F3'
    elif bld_no == 'B':
        if rm_no in flr_1:
            flr_no = 'SB2F1'
        elif rm_no in flr_2:
            flr_no = 'SB2F2'
        elif rm_no in flr_3:
            flr_no = 'SB2F3'
        elif rm_no in flr_4:
            flr_no = 'SB2F4'

    return flr_no


# @login_required
def write_to_excel_worksheet(request, excel_file, opbl):
    # BD_NO: 'A' AND  'B'
    excel_cell_dict = {

        'A': {'SB1F1': {'101': ['R12', 'H15', 'I15', 'J15', 'K15', 'L15', 'M15', 'N15', 'O15', 'P15', 'Q15', 'R15'],
                        '102': ['R27', 'H30', 'I30', 'J30', 'K30', 'L30', 'M30', 'N30', 'O30', 'P30', 'Q30', 'R30'],
                        '103': ['R42', 'H45', 'I45', 'J45', 'K45', 'L45', 'M45', 'N45', 'O45', 'P45', 'Q45', 'R45'],
                        '104': ['AG12', 'V15', 'W15', 'X15', 'Y15', 'Z15', 'AA15', 'AB15', 'AC15', 'AD15', 'AE15',
                                'AF15'],
                        '105': ['AG27', 'V30', 'W30', 'X30', 'Y30', 'Z30', 'AA30', 'AB30', 'AC30', 'AD30', 'AE30',
                                'AF30'],
                        '106': ['AG42', 'V45', 'W45', 'X45', 'Y45', 'Z45', 'AA45', 'AB45', 'AC45', 'AD45', 'AE45',
                                'AF45']},

              'SB1F2': {'201': ['R12', 'H15', 'I15', 'J15', 'K15', 'L15', 'M15', 'N15', 'O15', 'P15', 'Q15', 'R15'],
                        '202': ['R27', 'H30', 'I30', 'J30', 'K30', 'L30', 'M30', 'N30', 'O30', 'P30', 'Q30', 'R30'],
                        '203': ['R42', 'H45', 'I45', 'J45', 'K45', 'L45', 'M45', 'N45', 'O45', 'P45', 'Q45', 'R45'],
                        '204': ['AG12', 'V15', 'W15', 'X15', 'Y15', 'Z15', 'AA15', 'AB15', 'AC15', 'AD15', 'AE15',
                                'AF15'],
                        '205': ['AG27', 'V30', 'W30', 'X30', 'Y30', 'Z30', 'AA30', 'AB30', 'AC30', 'AD30', 'AE30',
                                'AF30'],
                        '206': ['AG42', 'V45', 'W45', 'X45', 'Y45', 'Z45', 'AA45', 'AB45', 'AC45', 'AD45', 'AE45',
                                'AF45']},

              'SB1F3': {'301': ['R12', 'H15', 'I15', 'J15', 'K15', 'L15', 'M15', 'N15', 'O15', 'P15', 'Q15', 'R15'],
                        '302': ['R27', 'H30', 'I30', 'J30', 'K30', 'L30', 'M30', 'N30', 'O30', 'P30', 'Q30', 'R30'],
                        '303': ['R42', 'H45', 'I45', 'J45', 'K45', 'L45', 'M45', 'N45', 'O45', 'P45', 'Q45', 'R45'],
                        '304': ['AG12', 'V15', 'W15', 'X15', 'Y15', 'Z15', 'AA15', 'AB15', 'AC15', 'AD15', 'AE15',
                                'AF15'],
                        '305': ['AG27', 'V30', 'W30', 'X30', 'Y30', 'Z30', 'AA30', 'AB30', 'AC30', 'AD30', 'AE30',
                                'AF30'],
                        '306': ['AG42', 'V45', 'W45', 'X45', 'Y45', 'Z45', 'AA45', 'AB45', 'AC45', 'AD45', 'AE45',
                                'AF45']}
              },

        'B': {'SB2F1': {'101': ['R12', 'H15', 'I15', 'J15', 'K15', 'L15', 'M15', 'N15', 'O15', 'P15', 'Q15', 'R15'],
                        '102': ['R27', 'H30', 'I30', 'J30', 'K30', 'L30', 'M30', 'N30', 'O30', 'P30', 'Q30', 'R30'],
                        '103': ['R42', 'H45', 'I45', 'J45', 'K45', 'L45', 'M45', 'N45', 'O45', 'P45', 'Q45', 'R45'],
                        '104': ['AG12', 'V15', 'W15', 'X15', 'Y15', 'Z15', 'AA15', 'AB15', 'AC15', 'AD15', 'AE15',
                                'AF15'],
                        '105': ['AG27', 'V30', 'W30', 'X30', 'Y30', 'Z30', 'AA30', 'AB30', 'AC30', 'AD30', 'AE30',
                                'AF30'],
                        '106': ['AG42', 'V45', 'W45', 'X45', 'Y45', 'Z45', 'AA45', 'AB45', 'AC45', 'AD45', 'AE45',
                                'AF45']},

              'SB2F2': {'201': ['R12', 'H15', 'I15', 'J15', 'K15', 'L15', 'M15', 'N15', 'O15', 'P15', 'Q15', 'R15'],
                        '202': ['R27', 'H30', 'I30', 'J30', 'K30', 'L30', 'M30', 'N30', 'O30', 'P30', 'Q30', 'R30'],
                        '203': ['R42', 'H45', 'I45', 'J45', 'K45', 'L45', 'M45', 'N45', 'O45', 'P45', 'Q45', 'R45'],
                        '204': ['AG12', 'V15', 'W15', 'X15', 'Y15', 'Z15', 'AA15', 'AB15', 'AC15', 'AD15', 'AE15',
                                'AF15'],
                        '205': ['AG27', 'V30', 'W30', 'X30', 'Y30', 'Z30', 'AA30', 'AB30', 'AC30', 'AD30', 'AE30',
                                'AF30'],
                        '206': ['AG42', 'V45', 'W45', 'X45', 'Y45', 'Z45', 'AA45', 'AB45', 'AC45', 'AD45', 'AE45',
                                'AF45']},

              'SB2F3': {'301': ['R12', 'H15', 'I15', 'J15', 'K15', 'L15', 'M15', 'N15', 'O15', 'P15', 'Q15', 'R15'],
                        '302': ['R27', 'H30', 'I30', 'J30', 'K30', 'L30', 'M30', 'N30', 'O30', 'P30', 'Q30', 'R30'],
                        '303': ['R42', 'H45', 'I45', 'J45', 'K45', 'L45', 'M45', 'N45', 'O45', 'P45', 'Q45', 'R45'],
                        '304': ['AG12', 'V15', 'W15', 'X15', 'Y15', 'Z15', 'AA15', 'AB15', 'AC15', 'AD15', 'AE15',
                                'AF15'],
                        '305': ['AG27', 'V30', 'W30', 'X30', 'Y30', 'Z30', 'AA30', 'AB30', 'AC30', 'AD30', 'AE30',
                                'AF30'],
                        '306': ['AG42', 'V45', 'W45', 'X45', 'Y45', 'Z45', 'AA45', 'AB45', 'AC45', 'AD45', 'AE45',
                                'AF45']},

              'SB2F4': {'401': ['R12', 'H15', 'I15', 'J15', 'K15', 'L15', 'M15', 'N15', 'O15', 'P15', 'Q15', 'R15'],
                        '402': ['R27', 'H30', 'I30', 'J30', 'K30', 'L30', 'M30', 'N30', 'O30', 'P30', 'Q30', 'R30'],
                        '403': ['R42', 'H45', 'I45', 'J45', 'K45', 'L45', 'M45', 'N45', 'O45', 'P45', 'Q45', 'R45'],
                        '404': ['AG12', 'V15', 'W15', 'X15', 'Y15', 'Z15', 'AA15', 'AB15', 'AC15', 'AD15', 'AE15',
                                'AF15'],
                        '405': ['AG27', 'V30', 'W30', 'X30', 'Y30', 'Z30', 'AA30', 'AB30', 'AC30', 'AD30', 'AE30',
                                'AF30'],
                        '406': ['AG42', 'V45', 'W45', 'X45', 'Y45', 'Z45', 'AA45', 'AB45', 'AC45', 'AD45', 'AE45',
                                'AF45']}
              },

    }

    # LOAD EXCEL WORKING FILE FROM DISK INTO PYTHON

    wb = load_workbook(excel_file)

    # CLEAR EXCEL WORKBOOK
    clear_excel_bill_slip_worksheet(wb, excel_cell_dict, excel_file)

    open_bill_date = opbl[0].bill_date  # from any open bill

    month = open_bill_date.month
    day = open_bill_date.day
    year = open_bill_date.year
    thai_year = get_thai_year(str(open_bill_date))

    for bill in opbl:
        r = bill.room_no
        fl_no = ''
        bd_no = r[3]
        rm_no = r[0:3]
        if fl_no != get_floor_no(r):
            fl_no = get_floor_no(r)

        # WRITE TO DATEBLOCK
        wb[fl_no]['D4'] = day
        wb[fl_no]['D5'] = month
        wb[fl_no]['D6'] = thai_year
        # WRITE TO EXCEL WORKSHEET
        cell = excel_cell_dict[bd_no][fl_no][rm_no]

        wb[fl_no][cell[0]] = bill.bill_ref
        wb[fl_no][cell[1]] = bill.tenant_name
        wb[fl_no][cell[2]] = bill.room_no[0:3]
        wb[fl_no][cell[3]] = bill.room_no[3]
        wb[fl_no][cell[4]] = bill.room_cost
        wb[fl_no][cell[5]] = bill.room_acc_cost
        wb[fl_no][cell[6]] = bill.electricity_cost
        wb[fl_no][cell[7]] = bill.water_cost
        wb[fl_no][cell[8]] = bill.common_ser_cost
        wb[fl_no][cell[9]] = bill.other_ser_cost
        wb[fl_no][cell[10]] = bill.adjust
        wb[fl_no][cell[11]] = bill.overdue_amount

    try:

        # SAVE TO DISK FILE
        cwd = os.getcwd()

        wb.save(excel_file)

        total_processed_rooms = len(opbl)

        messages.info(request,
                      'Bill Slip: {0}\\{1} with Total {2} Rooms have been SAVED !'.format(cwd, excel_file,
                                                                                          total_processed_rooms))

    except Exception as err:
        give_error_message('Error: '.format(err))


def clear_excel_bill_slip_worksheet(wb, worksheet_cell_dict, excel_file):
    # FULL OPEN-BILL-LIST (ALL ROOMS OCCUOIED)

    bd1_room = ['101A', '102A', '103A', '104A', '105A', '106A',
                '201A', '202A', '203A', '204A', '205A', '206A',
                '301A', '302A', '303A', '304A', '305A', '306A']

    bd2_room = ['101B', '102B', '103B', '104B', '105B', '106B',
                '201B', '202B', '203B', '204B', '205B', '206B',
                '301B', '302B', '303B', '304B', '305B', '306B',
                '401B', '402B', '403B', '404B', '405B', '406B']

    bd1_room.extend(bd2_room)  # COMBINE ALL ROOMS FROM BLD1 & 2 TO bd1_room

    for r in bd1_room:

        rm_no = r[0:3]  # 105
        bd_no = r[3]  # A
        floor_no = get_floor_no(r)  # r = 105A

        number_of_cell = 12  # NUMBER OF CELL to be cleared FOR EACH ROOM (EXCLUDE 'TOTAL')
        cell = worksheet_cell_dict[bd_no][floor_no][rm_no]  # LIST OF EXCEL CELLS

        # CLEAR EXCELL WORKSHEET

        # CLEAR DATE_BLOCK
        wb[floor_no]['D4'] = ''  # wb[tab]['D4']
        wb[floor_no]['D5'] = ''
        wb[floor_no]['D6'] = ''

        for i in range(number_of_cell):
            # CLEAR EXCEL-WORKSHEET
            wb[floor_no][cell[i]] = ''
    try:
        wb.save(excel_file)  # SAVE EXCEL WORKBOOK TO DISK FILE
    except Exception as err:
        give_error_message('Error: {}'.format(err))


# @login_required
def create_exel_sheet(request):
    os.chdir("c:\\users\\preechab\\dj_exel_file")
    excel_f = 'Month_Billing.xlsx'

    opbl = Billing.objects.filter(status='open').order_by('room_no')

    write_to_bill_summary(request, excel_f, opbl)
    write_to_excel_worksheet(request, excel_f, opbl)


@login_required
def tenant_comment(request):
    # ---TEST ONLY-------------
    # from pb_djams_project import settings

    # os.environ['SECRET_KEY'] = '75#^koi!0f__)fr2_3x#4qoatbv0wgp=+4(msc12!cav)gv@3&'
    # print(os.environ['SECRET_KEY'])

    # os.chdir("c:\\users\\preechab\\dj_exel_file")

    # excel_f = 'Month_Billing.xlsx'

    # cwd = os.getcwd()
    # bd = settings.BASE_DIR
    #
    # os.chdir(bd)
    #
    # messages.info(request, 'BASE_DIR: {}'.format(bd))
    # messages.info(request, 'CWD: {}'.format(cwd))
    # ---------------------------

    return render(request, 'ams/tenant_comment.html', {'section': 'comment'})


# @login_required
# def tenant_record(request):
#     return render(request, 'ams/tenant_record.html', {'section': 'record'})


def maintenance_charge(request):
    if request.method == 'POST':

        maintenance_form = MaintenanceForm(data=request.POST)

        if maintenance_form.is_valid():

            cd = maintenance_form.cleaned_data

            # Create a new object but avoid saving it yet
            new_ma_charge = maintenance_form.save(commit=False)

            new_ma_charge.desc = 'Maintenance cost'

            # Save the new object(MaintenanceCharge) to DB for ref.
            new_ma_charge.save()

            rmn = cd['room_no']
            pf = get_object_or_404(TenantProfile, room_no__room_no=rmn)

            # INCREAMENT & SAVE VALUE TO PF.MAINT_COST
            pf.maint_cost += cd['job_cost']
            pf.save()

            messages.success(request, 'Maintenance cost has been charged to Room: {}.'.format(rmn))

            return HttpResponseRedirect(reverse_lazy('admin_page'))
        else:
            messages.error(request, 'Error: new record was not saved !!!')

    else:
        maintenance_form = MaintenanceForm()

    return render(request, 'ams/maintenanace_charge.html', {'maintenance_form': maintenance_form})
